import calendar, random
from datetime import date
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

MONTHS={1:'ENERO',2:'FEBRERO',3:'MARZO',4:'ABRIL',5:'MAYO',6:'JUNIO',7:'JULIO',8:'AGOSTO',9:'SEPTIEMBRE',10:'OCTUBRE',11:'NOVIEMBRE',12:'DICIEMBRE'}
SUP=['ACCIETTO L.','ECHAVARRIA A.','AMAYA S.','MOCAYAR L.','MINGONE G.','GARAYZABAL D.','PEPA E.','DEVOTTO M.','RODRIGUEZ J.','MARTINEZ PAZ S.','DOMINGUEZ V.','BUSTOS FIERRO F.','JANISZEWSKI J.','MERLO C.','URZAGASTI F.']
ADMIN='CASTRO D.'; AGENTS=SUP+[ADMIN]; AERO=['AERO_01_07','AERO_07_13','AERO_13_19','AERO_19_01']
DEFAULT_LIC={'DOMINGUEZ V.':[2,3,4,5,6,7,8,9],'PEPA E.':[15,16,17,18,19,20,21,22,23],'RODRIGUEZ J.':list(range(1,15)),'URZAGASTI F.':[4,5],'GARAYZABAL D.':[13,14,15,16,17,18,19,20]}
DEFAULT_NO={'MERLO C.':[8,9,15,16,17,29,30],'DOMINGUEZ V.':[24,29],'JANISZEWSKI J.':[15,16,17],'URZAGASTI F.':[15,16,17]}

class GiroCustoms:
    def __init__(self,root):
        self.root=root;root.title('GIRO CUSTOMS — Cronograma y Reemplazos');root.geometry('1180x760');root.minsize(1000,680)
        self.r=None;self.replacement_log=[];self.build()
    def build(self):
        head=tk.Frame(self.root,bg='#1F4E79',height=80);head.pack(fill='x');tk.Label(head,text='GIRO CUSTOMS',bg='#1F4E79',fg='white',font=('Segoe UI',22,'bold')).pack(anchor='w',padx=28,pady=(10,0));tk.Label(head,text='Cronograma · Equidad · Ausencias · Reemplazo local',bg='#1F4E79',fg='#D9EAF7').pack(anchor='w',padx=30)
        nb=ttk.Notebook(self.root);nb.pack(fill='both',expand=True,padx=16,pady=12);self.tcfg=ttk.Frame(nb,padding=16);self.tabs=ttk.Frame(nb,padding=16);self.tprev=ttk.Frame(nb,padding=12);nb.add(self.tcfg,text='Configuración');nb.add(self.tabs,text='Ausencias → Reemplazo');nb.add(self.tprev,text='Cronograma')
        self.config_ui();self.abs_ui();self.preview_ui()
    def config_ui(self):
        f=ttk.LabelFrame(self.tcfg,text='Período y valores',padding=12);f.pack(fill='x');
        ttk.Label(f,text='Año').grid(row=0,column=0);self.year=ttk.Spinbox(f,from_=2020,to=2100,width=8);self.year.set(2026);self.year.grid(row=0,column=1,padx=8)
        ttk.Label(f,text='Mes').grid(row=0,column=2);self.month=ttk.Combobox(f,values=[f'{i:02d} - {MONTHS[i]}' for i in range(1,13)],state='readonly',width=18);self.month.current(7);self.month.grid(row=0,column=3,padx=8)
        ttk.Label(f,text='Feriados').grid(row=0,column=4);self.hol=ttk.Entry(f,width=14);self.hol.insert(0,'17');self.hol.grid(row=0,column=5,padx=8)
        ttk.Label(f,text='Hora 50%').grid(row=1,column=0,pady=8);self.v50=ttk.Entry(f,width=12);self.v50.insert(0,'17731.5');self.v50.grid(row=1,column=1)
        ttk.Label(f,text='Hora 100%').grid(row=1,column=2);self.v100=ttk.Entry(f,width=12);self.v100.insert(0,'23642');self.v100.grid(row=1,column=3)
        ttk.Label(f,text='Semilla').grid(row=1,column=4);self.seed=ttk.Entry(f,width=14);self.seed.grid(row=1,column=5)
        p=ttk.LabelFrame(self.tcfg,text='Restricciones',padding=12);p.pack(fill='both',expand=True,pady=12)
        a=ttk.Frame(p);a.pack(side='left',fill='both',expand=True,padx=(0,8));ttk.Label(a,text='LICENCIAS — NOMBRE: 1,2,3').pack(anchor='w');self.lic=tk.Text(a,font=('Consolas',9));self.lic.pack(fill='both',expand=True);self.lic.insert('1.0','\n'.join(f'{x}: {",".join(map(str,y))}' for x,y in DEFAULT_LIC.items()))
        b=ttk.Frame(p);b.pack(side='left',fill='both',expand=True,padx=(8,0));ttk.Label(b,text='NO TRABAJABLES — NOMBRE: 1,2,3').pack(anchor='w');self.no=tk.Text(b,font=('Consolas',9));self.no.pack(fill='both',expand=True);self.no.insert('1.0','\n'.join(f'{x}: {",".join(map(str,y))}' for x,y in DEFAULT_NO.items()))
        bar=ttk.Frame(self.tcfg);bar.pack(fill='x');self.status=tk.StringVar(value='Listo.');ttk.Label(bar,textvariable=self.status).pack(side='left');ttk.Button(bar,text='GENERAR CRONOGRAMA',command=self.generate).pack(side='right',padx=4);ttk.Button(bar,text='EXPORTAR EXCEL',command=self.export).pack(side='right',padx=4)
    def abs_ui(self):
        f=ttk.LabelFrame(self.tabs,text='Reemplazo automático de una ausencia',padding=12);f.pack(fill='x');ttk.Label(f,text='El mes ya generado queda congelado: solo se cambia el puesto afectado al confirmar.').pack(anchor='w')
        q=ttk.Frame(f);q.pack(fill='x',pady=10);ttk.Label(q,text='Día').pack(side='left');self.ad=ttk.Spinbox(q,from_=1,to=31,width=6);self.ad.set(1);self.ad.pack(side='left',padx=6);ttk.Label(q,text='Agente ausente').pack(side='left',padx=(18,4));self.aa=ttk.Combobox(q,values=AGENTS,state='readonly',width=24);self.aa.current(0);self.aa.pack(side='left');ttk.Label(q,text='Puesto/turno').pack(side='left',padx=(18,4));self.ak=ttk.Combobox(q,values=('TODOS LOS TURNOS',*AERO,*[f'ZONA_{i}' for i in range(1,7)]),state='readonly',width=22);self.ak.current(0);self.ak.pack(side='left');ttk.Button(q,text='BUSCAR REEMPLAZO',command=self.search).pack(side='left',padx=12);ttk.Button(q,text='HISTORIAL',command=self.history).pack(side='left')
        self.abs_status=tk.StringVar(value='');ttk.Label(f,textvariable=self.abs_status).pack(anchor='w')
        cols=('Día','Puesto','Turno','Original','Reemplazante','Índice','Estado');self.cv=ttk.Treeview(self.tabs,columns=cols,show='headings',height=14)
        for c,w in zip(cols,(55,150,100,165,170,270,110)):self.cv.heading(c,text=c);self.cv.column(c,width=w,anchor='center')
        self.cv.pack(fill='both',expand=True,pady=10)
    def preview_ui(self):
        self.pv=ttk.Treeview(self.tprev,columns=('Día','Agente','Sección','Zona','Turno','Pago'),show='headings')
        for c,w in zip(('Día','Agente','Sección','Zona','Turno','Pago'),(60,180,150,150,110,80)):self.pv.heading(c,text=c);self.pv.column(c,width=w,anchor='center')
        self.pv.pack(fill='both',expand=True)
    def parse(self,text):
        out={}
        for line in text.splitlines():
            if ':' not in line or not line.strip():continue
            name,vals=line.split(':',1);name=name.strip();out[name]=sorted({int(x.strip()) for x in vals.split(',') if x.strip()})
        return out
    def settings(self):
        y=int(self.year.get());m=self.month.current()+1;nd=calendar.monthrange(y,m)[1];hol={int(x.strip()) for x in self.hol.get().split(',') if x.strip()}
        if any(x<1 or x>nd for x in hol):raise ValueError('Hay un feriado fuera del mes seleccionado.')
        random.seed(int(self.seed.get().strip()) if self.seed.get().strip() else None)
        return y,m,nd,hol,float(self.v50.get().replace(',','.')),float(self.v100.get().replace(',','.')),self.parse(self.lic.get('1.0','end')),self.parse(self.no.get('1.0','end'))
    @staticmethod
    def ranges(v):
        if not v:return 'Ninguno'
        v=sorted(set(v));out=[];a=b=v[0]
        for x in v[1:]:
            if x==b+1:b=x
            else:out.append(str(a) if a==b else f'{a} al {b}');a=b=x
        out.append(str(a) if a==b else f'{a} al {b}');return ', '.join(out)
    def pay(self,r,d,t,aero):
        wd=date(r['year'],r['month'],d).weekday()
        if d in r['hol'] or wd==6:return '100%'
        if wd==5 and ((aero and t in ('13 A 19','19 A 01')) or (not aero and t=='15 A 22')):return '100%'
        return '50%'
    def stats(self,r):
        c={a:{k:0 for k in AERO+['SEC_50','SEC_100','TOTAL_TURNOS','H50','H100']} for a in AGENTS};zh={a:{z:0 for z in range(1,7)} for a in AGENTS};rows=[]
        for d,ass in r['cron'].items():
            for k,a in ass.items():
                if k.startswith('AERO_'):turn=k.replace('AERO_','').replace('_',' A ');sec='AEROPUERTO';zone='AEROPUERTO';h=6;p=self.pay(r,d,turn,True);c[a][k]+=1
                else:turn='15 A 22' if date(r['year'],r['month'],d).weekday()>=5 or d in r['hol'] else '19 A 02';sec='ZONA SECUNDARIA';zone=k.replace('ZONA_','ZONA ');h=7;p=self.pay(r,d,turn,False);c[a]['SEC_100' if p=='100%' else 'SEC_50']+=1;zh[a][int(k.split('_')[1])]+=1
                c[a]['TOTAL_TURNOS']+=1;c[a]['H100' if p=='100%' else 'H50']+=h;rows.append({'DÍA':d,'AGENTE':a,'SECCIÓN':sec,'ZONA':zone,'TURNO':turn,'PAGO':p})
        control=[]
        for a in AGENTS:
            x=c[a];v50=x['H50']*r['v50'];v100=x['H100']*r['v100'];tot=v50+v100;control.append({'Agente':a,'Dias_Disponibles':r['avail'][a],**{k:x[k] for k in AERO+['SEC_50','SEC_100','TOTAL_TURNOS']},'CANT_HORAS_50':x['H50'],'CANT_HORAS_100':x['H100'],'CANT_HORAS_TOTALES':x['H50']+x['H100'],'VALOR_TOTAL_50':v50,'VALOR_TOTAL_100':v100,'VALOR_TOTAL_AGENTE':tot,'Sueldo_Proporcional_Por_Dia':round(tot/r['avail'][a],2)})
        r['zone_hist']=zh;r['df_cron']=pd.DataFrame(rows);r['df_control']=pd.DataFrame(control);r['df_lic']=pd.DataFrame([{'Agente':a,'Días Licencia':len(r['blocked'][a]['lic']),'Fechas Licencia':self.ranges(r['blocked'][a]['lic']),'Días Extensión (Finde/Feriado)':len(r['blocked'][a]['ext']),'Fechas Extensión Finde/Feriado':self.ranges(r['blocked'][a]['ext']),'Días No Trabajables':len(r['blocked'][a]['no']),'Fechas Días No Trabajables':self.ranges(r['blocked'][a]['no']),'Total Días No Disponibles':len(r['blocked'][a]['todos'])} for a in AGENTS])
    def generate(self):
        try:
            y,m,nd,hol,v50,v100,lic,no=self.settings();week=lambda d:date(y,m,d).weekday()>=5 or d in hol;blocked={}
            for a in AGENTS:
                l=set(lic.get(a,[]));n=set(no.get(a,[]));e={d+1 for d in l if d<nd and week(d+1) and d+1 not in l};blocked[a]={'lic':l,'ext':e,'no':n,'todos':l|e|n}
            r={'year':y,'month':m,'days':nd,'hol':hol,'v50':v50,'v100':v100,'blocked':blocked,'avail':{a:max(1,nd-len(blocked[a]['todos'])) for a in AGENTS},'cron':{d:{} for d in range(1,nd+1)},'zone_hist':{a:{z:0 for z in range(1,7)} for a in AGENTS}}
            zone_castro=1
            for d in range(1,nd+1):
                late=set()
                if d>1:
                    late|={r['cron'][d-1].get('AERO_19_01')};late|={a for k,a in r['cron'][d-1].items() if k.startswith('ZONA_')}
                active=('01 A 07','07 A 13','13 A 19','19 A 01') if week(d) else ('01 A 07','19 A 01')
                for t in active:
                    key='AERO_'+t.replace(' A ','_');self.stats(r);df=r['df_control'];cand=[a for a in SUP if d not in blocked[a]['todos'] and a not in r['cron'][d].values()]
                    if d>1:cand=[a for a in cand if a not in {r['cron'][d-1].get(k) for k in AERO}]
                    if t=='01 A 07':
                        q=[a for a in cand if a not in late]
                        if q:cand=q
                    cand.sort(key=lambda a:(df.loc[df.Agente==a,key].iloc[0],df.loc[df.Agente==a,'TOTAL_TURNOS'].iloc[0]/r['avail'][a],df.loc[df.Agente==a,'CANT_HORAS_TOTALES'].iloc[0],df.loc[df.Agente==a,'VALOR_TOTAL_AGENTE'].iloc[0]/r['avail'][a]))
                    if cand:r['cron'][d][key]=cand[0]
                if d not in blocked[ADMIN]['todos']:r['cron'][d][f'ZONA_{zone_castro}']=ADMIN;zone_castro=zone_castro+1 if zone_castro<6 else 1
                occupied=zone_castro-1 if zone_castro>1 else 6;zones=[1,2,3,4,5,6] if d in blocked[ADMIN]['todos'] else [z for z in range(1,7) if z!=occupied]
                for z in zones:
                    self.stats(r);df=r['df_control'];cand=[a for a in AGENTS if d not in blocked[a]['todos'] and a not in r['cron'][d].values()];cand.sort(key=lambda a:(df.loc[df.Agente==a,'VALOR_TOTAL_AGENTE'].iloc[0]/r['avail'][a],df.loc[df.Agente==a,'TOTAL_TURNOS'].iloc[0]/r['avail'][a],r['zone_hist'][a][z]))
                    if cand:r['cron'][d][f'ZONA_{z}']=cand[0]
            self.stats(r);self.r=r;self.replacement_log=[];self.refresh();self.status.set(f'Generado {MONTHS[m]} {y}.');messagebox.showinfo('GIRO CUSTOMS','Cronograma generado correctamente.')
        except Exception as e:messagebox.showerror('Error de generación',str(e))
    def refresh(self):
        for i in self.pv.get_children():self.pv.delete(i)
        if self.r:
            for x in self.r['df_cron'].itertuples(index=False):self.pv.insert('', 'end',values=tuple(x))
    def candidate_rank(self,day,key,absent,reserved):
        r=self.r;pool=SUP if key.startswith('AERO_') else AGENTS;turn=key.replace('AERO_','').replace('_',' A ') if key.startswith('AERO_') else ('15 A 22' if date(r['year'],r['month'],day).weekday()>=5 or day in r['hol'] else '19 A 02');aero=key.startswith('AERO_');out=[]
        for a in pool:
            if a in reserved or a==absent or day in r['blocked'][a]['todos'] or a in r['cron'][day].values():continue
            prev=r['cron'].get(day-1,{}) if day>1 else {}
            if aero:
                if any(prev.get(k)==a for k in AERO):continue
                if key=='AERO_01_07' and (prev.get('AERO_19_01')==a or any(v==a for k,v in prev.items() if k.startswith('ZONA_'))):continue
            elif any(v==a for k,v in prev.items() if k.startswith('ZONA_')):continue
            row=r['df_control'][r['df_control'].Agente==a].iloc[0];exact=row[key] if key in row.index else row['SEC_100' if self.pay(r,day,turn,aero)=='100%' else 'SEC_50'];zone=r['zone_hist'][a][int(key.split('_')[1])] if key.startswith('ZONA_') else 0;score=(exact,row.TOTAL_TURNOS/r['avail'][a],row.CANT_HORAS_TOTALES/r['avail'][a],row.VALOR_TOTAL_AGENTE/r['avail'][a],zone);out.append((score,a,turn,self.pay(r,day,turn,aero)))
        return sorted(out,key=lambda x:(x[0],x[1]))
    def search(self):
        if not self.r:return messagebox.showwarning('Sin cronograma','Primero genera el cronograma.')
        try:d=int(self.ad.get())
        except ValueError:return messagebox.showerror('Dato inválido','Día inválido.')
        if d<1 or d>self.r['days']:return messagebox.showerror('Dato inválido','El día no pertenece al mes.')
        absent=self.aa.get();sel=self.ak.get();keys=[k for k,v in self.r['cron'][d].items() if v==absent] if sel=='TODOS LOS TURNOS' else [sel]
        for i in self.cv.get_children():self.cv.delete(i)
        if not keys:return messagebox.showinfo('Sin asignación','El agente no tiene ese puesto/turno ese día.')
        reserved=set();self.pending=[]
        for key in keys:
            if self.r['cron'][d].get(key)!=absent:continue
            ranked=self.candidate_rank(d,key,absent,reserved)
            if not ranked:continue
            best=ranked[0];reserved.add(best[1]);self.pending.append((d,key,absent,best[1],best[2],best[3],best[0]))
            for i,x in enumerate(ranked[:8]):self.cv.insert('', 'end',values=(d,key,x[2],absent,x[1],str(x[0]),'MEJOR' if i==0 else 'ALTERNATIVA'))
        if not self.pending:messagebox.showwarning('Sin reemplazo','No existe candidato compatible con las restricciones.')
        else:self.abs_status.set('Propuesta preparada. Confirme solo después de revisar el candidato.')
    def confirm(self):
        if not self.pending:return messagebox.showwarning('Sin propuesta','Pulse BUSCAR REEMPLAZO primero.')
        if not messagebox.askyesno('Confirmar','Se modificarán únicamente los puestos indicados. ¿Continuar?'):return
        changes=[]
        for d,key,old,new,turn,pay,score in self.pending:
            if self.r['cron'][d].get(key)==old:
                self.r['cron'][d][key]=new;changes.append((d,key,old,new,turn,pay))
        if not changes:return messagebox.showwarning('No aplicado','La asignación cambió antes de confirmar.')
        self.stats(self.r)
        for d,key,old,new,turn,pay in changes:self.replacement_log.append({'DÍA':d,'PUESTO':key,'ORIGINAL':old,'REEMPLAZO':new,'TURNO':turn,'PAGO':pay})
        self.pending=[];self.refresh();self.abs_status.set('Reemplazo confirmado. Ningún otro puesto fue regenerado.');messagebox.showinfo('Confirmado','Reemplazo aplicado correctamente.')
    def history(self):
        w=tk.Toplevel(self.root);w.title('Historial de reemplazos');w.geometry('800x400');cols=('DÍA','PUESTO','ORIGINAL','REEMPLAZO','TURNO','PAGO');t=ttk.Treeview(w,columns=cols,show='headings')
        for c in cols:t.heading(c,text=c);t.column(c,width=125,anchor='center')
        for x in self.replacement_log:t.insert('', 'end',values=tuple(x[c] for c in cols))
        t.pack(fill='both',expand=True,padx=10,pady=10)
    def export(self):
        if not self.r:return messagebox.showwarning('Sin datos','Genere primero el cronograma.')
        path=filedialog.asksaveasfilename(defaultextension='.xlsx',initialfile=f'Giro_Customs_{MONTHS[self.r["month"]]}_{self.r["year"]}.xlsx',filetypes=[('Excel','*.xlsx')])
        if not path:return
        try:
            r=self.r;wb=openpyxl.Workbook();ws=wb.active;ws.title='Cronograma Día a Día';mat=wb.create_sheet('Cronograma Consolidado');ct=wb.create_sheet('Control de Equidad');li=wb.create_sheet('Consolidado de Licencias');blue=PatternFill('solid',fgColor='1F4E79');hf=Font(name='Segoe UI',size=10,bold=True,color='FFFFFF');fd=Font(name='Segoe UI',size=9);bd=Border(*(Side('thin',color='BDC3C7'),)*4)
            ws['A1']=f'CRONOGRAMA DE ASIGNACIONES - {MONTHS[r["month"]]} {r["year"]}';self.sheet_df(ws,r['df_cron'],blue,hf,fd,bd)
            mat.merge_cells('B3:C3');mat['B3']='DIA - HORARIO A CUBRIR';mat.merge_cells('D3:G3');mat['D3']='SECCIÓN AEROPUERTO';mat.merge_cells('H3:M3');mat['H3']='ZONA SECUNDARIA';heads=['01 A 07','07 A 13','13 A 19','19 A 01','ZONA 1','ZONA 2','ZONA 3','ZONA 4','ZONA 5','ZONA 6']
            for i,h in enumerate(heads,4):mat.cell(4,i,h);mat.cell(4,i).fill=blue;mat.cell(4,i).font=hf;mat.cell(4,i).alignment=Alignment(horizontal='center')
            names=['LUNES','MARTES','MIÉRCOLES','JUEVES','VIERNES','SÁBADO','DOMINGO'];mapping=AERO+[f'ZONA_{i}' for i in range(1,7)]
            for d in range(1,r['days']+1):
                rr=d+4;mat.cell(rr,2,names[date(r['year'],r['month'],d).weekday()]);mat.cell(rr,3,d)
                for ci,k in enumerate(mapping,4):mat.cell(rr,ci,r['cron'][d].get(k,''));mat.cell(rr,ci).border=bd;mat.cell(rr,ci).alignment=Alignment(horizontal='center')
            ct['A1']=f'MÉTRICAS Y CONTROL DE EQUIDAD - {MONTHS[r["month"]]} {r["year"]}';self.sheet_df(ct,r['df_control'],blue,hf,fd,bd,True);li['A1']=f'CONSOLIDADO DE LICENCIAS Y DÍAS NO TRABAJABLES - {MONTHS[r["month"]]} {r["year"]}';self.sheet_df(li,r['df_lic'],blue,hf,fd,bd)
            if self.replacement_log:self.sheet_df(wb.create_sheet('Historial Reemplazos'),pd.DataFrame(self.replacement_log),blue,hf,fd,bd)
            for sh in wb.worksheets:
                for col in sh.columns:
                    letter=get_column_letter(col[0].column);sh.column_dimensions[letter].width=max(12,min(42,max((len(str(c.value or '')) for c in col),default=10)+3))
            wb.save(path);self.status.set(f'Excel guardado: {path}');messagebox.showinfo('Excel','Archivo exportado correctamente.')
        except Exception as e:messagebox.showerror('Error Excel',str(e))
    @staticmethod
    def sheet_df(ws,df,fill,hf,fd,bd,currency=False):
        for ci,t in enumerate(df.columns,1):c=ws.cell(3,ci,t);c.fill=fill;c.font=hf;c.alignment=Alignment(horizontal='center')
        for ri,row in enumerate(df.itertuples(index=False),4):
            for ci,v in enumerate(row,1):c=ws.cell(ri,ci,v);c.font=fd;c.border=bd;c.alignment=Alignment(horizontal='center' if ci>1 else 'left')
            if currency:
                for c in ws[ri][12:]:c.number_format='$#,##0.00'
if __name__=='__main__':
    root=tk.Tk();GiroCustoms(root);root.mainloop()
