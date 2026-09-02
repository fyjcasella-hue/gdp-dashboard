import calendar
import os
import tkinter as tk
from tkinter import ttk
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

import giro_supervisores_launcher_v4 as base
from giro_supervisores_launcher_v4 import GiroApp


def _show_calendar_holiday_selectable(self):
    if not hasattr(self, 'cal_frame'):
        return
    self.parse_month()
    for w in self.cal_frame.winfo_children():
        w.destroy()
    agent = self.cal_agent.get() if hasattr(self, 'cal_agent') else None
    if not agent:
        return
    mode = self.cal_mode.get()
    licenses = self.licenses.setdefault(agent, set())
    unavailable = self.unavailable.setdefault(agent, set())
    self.calendar_agent = agent
    heads = ['L', 'M', 'X', 'J', 'V', 'S', 'D']
    for c, h in enumerate(heads):
        ttk.Label(self.cal_frame, text=h, font=('Segoe UI', 10, 'bold'), anchor='center').grid(row=0, column=c, sticky='nsew', padx=3, pady=3)
    first, days = calendar.monthrange(self.year, self.month)
    for i in range(first + days):
        d = i - first + 1
        if d < 1:
            continue
        is_holiday = d in self.holidays
        is_license = d in licenses
        is_unavailable = d in unavailable
        if is_license:
            bg = '#F7B7B7'
        elif is_unavailable:
            bg = '#C9D7FF'
        elif is_holiday:
            bg = '#FFE3A3'
        else:
            bg = '#F3F5F7'
        text = str(d)
        if is_holiday and (is_license or is_unavailable):
            text += '\nFERIADO'
        r = i // 7 + 1
        c = i % 7
        tk.Button(self.cal_frame, text=text, bg=bg, activebackground=bg, relief='flat', font=('Segoe UI', 11, 'bold'), width=8, height=3, command=lambda x=d: self.toggle_day(x)).grid(row=r, column=c, sticky='nsew', padx=4, pady=4)
    for c in range(7):
        self.cal_frame.columnconfigure(c, weight=1)
    self.cal_help.set(f'{agent}: modo {mode}. Rojo = licencia · Azul = no disponible · Amarillo = feriado. Un feriado también puede marcarse como licencia o no disponible.')


def _toggle_day_holiday_selectable(self, d):
    agent = self.cal_agent.get()
    if not agent:
        return
    target = self.licenses.setdefault(agent, set()) if self.cal_mode.get() == 'LICENCIA' else self.unavailable.setdefault(agent, set())
    if d in target:
        target.remove(d)
    else:
        target.add(d)
    self.show_calendar()


GiroApp.show_calendar = _show_calendar_holiday_selectable
GiroApp.refresh_calendar = _show_calendar_holiday_selectable
GiroApp.toggle_day = _toggle_day_holiday_selectable

# Primero se ejecuta la exportación existente; después se agrega únicamente
# la solapa que faltaba, conservando intactas las demás hojas.
_original_export_excel = GiroApp.export_excel


def _export_excel_with_consolidated(self):
    captured = {'path': None}
    original_dialog = base.filedialog.asksaveasfilename

    def _capture_dialog(*args, **kwargs):
        path = original_dialog(*args, **kwargs)
        captured['path'] = path
        return path

    base.filedialog.asksaveasfilename = _capture_dialog
    try:
        _original_export_excel(self)
    finally:
        base.filedialog.asksaveasfilename = original_dialog

    path = captured['path']
    if not path or not os.path.exists(path) or not self.result:
        return

    wb = load_workbook(path)
    if 'Cronograma Consolidado' in wb.sheetnames:
        del wb['Cronograma Consolidado']
    ws = wb.create_sheet('Cronograma Consolidado', 1)
    ws.sheet_view.showGridLines = True

    header_fill = PatternFill('solid', fgColor='1F4E79')
    day_fill = PatternFill('solid', fgColor='D9E1F2')
    weekend_fill = PatternFill('solid', fgColor='FEF2CB')
    weekday_aero_fill = PatternFill('solid', fgColor='E2EFDA')
    zebra_fill = PatternFill('solid', fgColor='F9FAFB')
    white_fill = PatternFill('solid', fgColor='FFFFFF')
    thin = Side(style='thin', color='BDC3C7')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    font_header = Font(name='Segoe UI', size=10, bold=True, color='FFFFFF')
    font_data = Font(name='Segoe UI', size=9)
    font_bold = Font(name='Segoe UI', size=9, bold=True)
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center')

    ws.merge_cells('B3:C3'); ws['B3'] = 'DIA - HORARIO A CUBRIR'
    ws.merge_cells('D3:G3'); ws['D3'] = 'SECCIÓN AEROPUERTO'
    ws.merge_cells('H3:M3'); ws['H3'] = 'ZONA SECUNDARIA'
    for ref in ('B3', 'D3', 'H3'):
        ws[ref].fill = header_fill; ws[ref].font = font_header; ws[ref].alignment = center; ws[ref].border = border

    headers = [(4,'01 A 07'),(5,'07 A 13'),(6,'13 A 19'),(7,'19 A 01'),(8,'ZONA 1'),(9,'ZONA 2'),(10,'ZONA 3'),(11,'ZONA 4'),(12,'ZONA 5'),(13,'ZONA 6')]
    for col, text in headers:
        cell = ws.cell(row=4, column=col, value=text); cell.fill = header_fill; cell.font = font_header; cell.alignment = center; cell.border = border

    weekdays = ['LUNES','MARTES','MIÉRCOLES','JUEVES','VIERNES','SÁBADO','DOMINGO']
    ndays = calendar.monthrange(self.year, self.month)[1]
    cron = self.result['cron']
    mapping = [(4,'AERO_01_07'),(5,'AERO_07_13'),(6,'AERO_13_19'),(7,'AERO_19_01'),(8,'ZONA_1'),(9,'ZONA_2'),(10,'ZONA_3'),(11,'ZONA_4'),(12,'ZONA_5'),(13,'ZONA_6')]

    for d in range(1, ndays + 1):
        row = d + 4
        weekday = weekdays[calendar.weekday(self.year, self.month, d)]
        special = d in self.holidays or calendar.weekday(self.year, self.month, d) in (5,6)
        c = ws.cell(row=row, column=2, value=weekday); c.fill=day_fill; c.font=font_bold; c.border=border; c.alignment=left
        c = ws.cell(row=row, column=3, value=d); c.fill=day_fill; c.font=font_bold; c.border=border; c.alignment=center
        for col, key in mapping:
            agent = cron.get(d, {}).get(key, '')
            c = ws.cell(row=row, column=col, value=agent); c.font=font_data; c.border=border; c.alignment=center
            if special: c.fill=weekend_fill
            elif col in (4,7): c.fill=weekday_aero_fill
            else: c.fill=white_fill if row % 2 == 0 else zebra_fill

    ws.column_dimensions['A'].width=3; ws.column_dimensions['B'].width=14; ws.column_dimensions['C'].width=6
    for col in range(4,14): ws.column_dimensions[get_column_letter(col)].width=20
    wb.save(path)


GiroApp.export_excel = _export_excel_with_consolidated


if __name__ == '__main__':
    root = tk.Tk()
    GiroApp(root)
    root.mainloop()
