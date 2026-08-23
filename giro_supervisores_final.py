import calendar
import random
from datetime import date
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

APP_NAME = 'GIRO DE SUPERVISORES - NELSON CASELLA'
MONTHS = ['ENERO','FEBRERO','MARZO','ABRIL','MAYO','JUNIO','JULIO','AGOSTO','SEPTIEMBRE','OCTUBRE','NOVIEMBRE','DICIEMBRE']
DEFAULT_SUP = ['ACCIETTO L.','ECHAVARRIA A.','AMAYA S.','MOCAYAR L.','MINGONE G.','GARAYZABAL D.','PEPA E.','DEVOTTO M.','RODRIGUEZ J.','MARTINEZ PAZ S.','DOMINGUEZ V.','BUSTOS FIERRO F.','JANISZEWSKI J.','MERLO C.','URZAGASTI F.']
DEFAULT_ADMIN = 'CASTRO D.'
AERO = ['AERO_01_07','AERO_07_13','AERO_13_19','AERO_19_01']

class GiroApp:
    def __init__(self, root):
        self.root = root
        self.root.title(APP_NAME)
        self.root.geometry('1280x820')
        self.root.minsize(1100, 720)
        self.sup = list(DEFAULT_SUP)
        self.admin = DEFAULT_ADMIN
        self.active_agents = set(self.sup + [self.admin])
        self.licenses = {a:set() for a in self.sup + [self.admin]}
        self.unavailable = {a:set() for a in self.sup + [self.admin]}
        self.holidays = set()
        self.year = 2026
        self.month = 8
        self.v50 = 17731.5
        self.v100 = 23642.0
        self.result = None
        self.pending = []
        self.replacement_log = []
        self.calendar_mode = 'LICENCIA'
        self.calendar_agent = None
        self.build_style()
        self.build_ui()
        self.refresh_agents()
        self.show_calendar()

    def build_style(self):
        s = ttk.Style()
        try: s.theme_use('clam')
        except tk.TclError: pass
        s.configure('TButton', font=('Segoe UI',10), padding=(12,7))
        s.configure('Accent.TButton', font=('Segoe UI',10,'bold'), padding=(14,8))
        s.configure('TLabel', font=('Segoe UI',10))
        s.configure('Title.TLabel', font=('Segoe UI',20,'bold'))
        s.configure('Sub.TLabel', font=('Segoe UI',10))
        s.configure('Treeview', rowheight=28, font=('Segoe UI',9))
        s.configure('Treeview.Heading', font=('Segoe UI',9,'bold'))
        s.configure('TNotebook.Tab', font=('Segoe UI',10,'bold'), padding=(15,8))

    def build_ui(self):
        header = tk.Frame(self.root, bg='#172B4D', height=92)
        header.pack(fill='x')
        tk.Label(header, text='GIRO DE SUPERVISORES', bg='#172B4D', fg='white', font=('Segoe UI',24,'bold')).pack(anchor='w', padx=28, pady=(13,0))
        tk.Label(header, text='Nelson Casella  |  Cronograma inteligente · Equidad · Licencias · Ausencias', bg='#172B4D', fg='#B8C7DB', font=('Segoe UI',10)).pack(anchor='w', padx=30)
        self.nb = ttk.Notebook(self.root)
        self.nb.pack(fill='both', expand=True, padx=14, pady=14)
        self.tab_config = ttk.Frame(self.nb, padding=16)
        self.tab_agents = ttk.Frame(self.nb, padding=16)
        self.tab_calendar = ttk.Frame(self.nb, padding=16)
        self.tab_abs = ttk.Frame(self.nb, padding=16)
        self.tab_schedule = ttk.Frame(self.nb, padding=12)
        self.nb.add(self.tab_config, text='Configuración')
        self.nb.add(self.tab_agents, text='Agentes')
        self.nb.add(self.tab_calendar, text='Calendario')
        self.nb.add(self.tab_abs, text='Ausencias → Reemplazo')
        self.nb.add(self.tab_schedule, text='Cronograma')
        self.config_tab(); self.agents_tab(); self.calendar_tab(); self.absence_tab(); self.schedule_tab()
        self.status = tk.StringVar(value='Listo.')
        ttk.Label(self.root, textvariable=self.status, anchor='w').pack(fill='x', padx=18, pady=(0,8))

    def config_tab(self):
        box = ttk.LabelFrame(self.tab_config, text='Período y valores', padding=14); box.pack(fill='x')
        ttk.Label(box,text='Año').grid(row=0,column=0,sticky='w'); self.year_var=tk.IntVar(value=2026); ttk.Spinbox(box,from_=2020,to=2100,textvariable=self.year_var,width=8,command=self.refresh_calendar).grid(row=0,column=1,padx=8)
        ttk.Label(box,text='Mes').grid(row=0,column=2,sticky='w'); self.month_var=tk.StringVar(value='08 - AGOSTO'); self.month_cb=ttk.Combobox(box,textvariable=self.month_var,state='readonly',values=[f'{i:02d} - {MONTHS[i-1]}' for i in range(1,13)],width=18); self.month_cb.grid(row=0,column=3,padx=8); self.month_cb.bind('<<ComboboxSelected>>',lambda e:self.refresh_calendar())
        ttk.Label(box,text='Valor hora 50%').grid(row=1,column=0,pady=10,sticky='w'); self.v50_var=tk.StringVar(value='17731.5'); ttk.Entry(box,textvariable=self.v50_var,width=12).grid(row=1,column=1,padx=8)
        ttk.Label(box,text='Valor hora 100%').grid(row=1,column=2,sticky='w'); self.v100_var=tk.StringVar(value='23642'); ttk.Entry(box,textvariable=self.v100_var,width=12).grid(row=1,column=3,padx=8)
        ttk.Label(box,text='Feriados del mes').grid(row=0,column=4,padx=(25,0),sticky='w'); self.holiday_var=tk.StringVar(value='17'); ttk.Entry(box,textvariable=self.holiday_var,width=18).grid(row=0,column=5,padx=8)
        ttk.Button(box,text='APLICAR FERIADOS',command=self.apply_holidays).grid(row=1,column=5,padx=8,sticky='e')
        info=ttk.LabelFrame(self.tab_config,text='Regla de continuidad entre meses',padding=14); info.pack(fill='x',pady=14)
        ttk.Label(info,text='El día 1 de cada mes NO hereda restricciones de turnos del último día del mes anterior. El nuevo mes comienza desde cero.',font=('Segoe UI',10,'bold')).pack(anchor='w')
        ttk.Label(info,text='Sí se respetan todas las licencias, días no disponibles, feriados y reglas de descanso dentro del nuevo mes.').pack(anchor='w',pady=(5,0))
        actions=ttk.Frame(self.tab_config); actions.pack(fill='x',pady=10)
        ttk.Button(actions,text='GENERAR CRONOGRAMA',style='Accent.TButton',command=self.generate).pack(side='left')
        ttk.Button(actions,text='EXPORTAR EXCEL',command=self.export_excel).pack(side='left',padx=8)

    def agents_tab(self):
        left=ttk.LabelFrame(self.tab_agents,text='Agentes activos',padding=10); left.pack(side='left',fill='both',expand=True,padx=(0,8))
        self.agent_tree=ttk.Treeview(left,columns=('Nombre','Tipo','Estado'),show='headings',height=20)
        for c,w in [('Nombre',230),('Tipo',180),('Estado',100)]: self.agent_tree.heading(c,text=c); self.agent_tree.column(c,width=w,anchor='center')
        self.agent_tree.pack(fill='both',expand=True)
        right=ttk.LabelFrame(self.tab_agents,text='Alta / baja',padding=14); right.pack(side='left',fill='y',padx=(8,0))
        ttk.Label(right,text='Nombre completo / identificación').pack(anchor='w'); self.new_agent=tk.StringVar(); ttk.Entry(right,textvariable=self.new_agent,width=28).pack(pady=(4,12))
        ttk.Label(right,text='Tipo').pack(anchor='w'); self.agent_type=tk.StringVar(value='SUPERVISOR AEROPUERTO'); ttk.Combobox(right,textvariable=self.agent_type,state='readonly',values=['SUPERVISOR AEROPUERTO','ADMINISTRADOR FIJO'],width=25).pack(pady=(4,16))
        ttk.Button(right,text='DAR DE ALTA',style='Accent.TButton',command=self.add_agent).pack(fill='x',pady=4)
        ttk.Button(right,text='DAR DE BAJA',command=self.remove_agent).pack(fill='x',pady=4)
        ttk.Label(right,text='La baja no borra información histórica ni altera cronogramas ya generados.',wraplength=260).pack(anchor='w',pady=20)

    def calendar_tab(self):
        top=ttk.Frame(self.tab_calendar); top.pack(fill='x')
        ttk.Label(top,text='Agente').pack(side='left'); self.cal_agent=ttk.Combobox(top,state='readonly',width=28); self.cal_agent.pack(side='left',padx=8); self.cal_agent.bind('<<ComboboxSelected>>',lambda e:self.show_calendar())
        self.cal_mode=tk.StringVar(value='LICENCIA')
        for mode in ['LICENCIA','NO DISPONIBLE']:
            ttk.Radiobutton(top,text=mode,value=mode,variable=self.cal_mode,command=self.show_calendar).pack(side='left',padx=6)
        ttk.Button(top,text='LIMPIAR MES',command=self.clear_agent_month).pack(side='right')
        self.cal_frame=ttk.Frame(self.tab_calendar); self.cal_frame.pack(fill='both',expand=True,pady=15)
        self.cal_help=tk.StringVar(value='Hacé clic sobre los días para marcarlos.')
        ttk.Label(self.tab_calendar,textvariable=self.cal_help).pack(anchor='w')

    def absence_tab(self):
        box=ttk.LabelFrame(self.tab_abs,text='Ausencia puntual — reparación local',padding=14); box.pack(fill='x')
        ttk.Label(box,text='El cronograma existente NO se regenera. Solo se reemplaza el puesto afectado.').pack(anchor='w')
        row=ttk.Frame(box); row.pack(fill='x',pady=12)
        ttk.Label(row,text='Día').pack(side='left'); self.abs_day=tk.IntVar(value=1); ttk.Spinbox(row,from_=1,to=31,textvariable=self.abs_day,width=6).pack(side='left',padx=6)
        ttk.Label(row,text='Agente ausente').pack(side='left',padx=(18,5)); self.abs_agent=ttk.Combobox(row,state='readonly',width=25); self.abs_agent.pack(side='left')
        ttk.Label(row,text='Puesto').pack(side='left',padx=(18,5)); self.abs_key=ttk.Combobox(row,state='readonly',width=23,values=['TODOS LOS TURNOS',*AERO,*[f'ZONA_{i}' for i in range(1,7)]]); self.abs_key.current(0); self.abs_key.pack(side='left')
        ttk.Button(row,text='BUSCAR REEMPLAZO',style='Accent.TButton',command=self.search_replacement).pack(side='left',padx=12)
        self.abs_msg=tk.StringVar(value='Generá primero el cronograma.'); ttk.Label(box,textvariable=self.abs_msg).pack(anchor='w')
        cols=('Día','Puesto','Turno','Original','Mejor candidato','Carga','Estado'); self.abs_tree=ttk.Treeview(self.tab_abs,columns=cols,show='headings',height=16)
        for c,w in zip(cols,[55,150,105,175,190,260,120]): self.abs_tree.heading(c,text=c); self.abs_tree.column(c,width=w,anchor='center')
        self.abs_tree.pack(fill='both',expand=True,pady=12)
        ttk.Button(self.tab_abs,text='CONFIRMAR REEMPLAZO SELECCIONADO',command=self.confirm_replacement).pack(side='right')
        ttk.Button(self.tab_abs,text='VER HISTORIAL',command=self.show_history).pack(side='right',padx=8)

    def schedule_tab(self):
        top=ttk.Frame(self.tab_schedule); top.pack(fill='x'); ttk.Button(top,text='ACTUALIZAR VISTA',command=self.refresh_schedule).pack(side='left'); ttk.Button(top,text='EXPORTAR EXCEL',command=self.export_excel).pack(side='right')
        cols=('DÍA','AGENTE','SECCIÓN','ZONA','TURNO','PAGO'); self.schedule_tree=ttk.Treeview(self.tab_schedule,columns=cols,show='headings')
        for c,w in zip(cols,[60,200,170,150,110,90]): self.schedule_tree.heading(c,text=c); self.schedule_tree.column(c,width=w,anchor='center')
        self.schedule_tree.pack(fill='both',expand=True,pady=10)

    def all_agents(self): return self.sup + [self.admin] if self.admin in self.active_agents else list(self.sup)
    def refresh_agents(self):
        active=self.all_agents()
        for tree in [self.agent_tree]:
            for x in tree.get_children(): tree.delete(x)
        for a in active:
            typ='ADMINISTRADOR FIJO' if a==self.admin else 'SUPERVISOR AEROPUERTO'; self.agent_tree.insert('', 'end', values=(a,typ,'ACTIVO'))
        self.cal_agent['values']=active; self.abs_agent['values']=active
        if active and self.cal_agent.get() not in active: self.cal_agent.current(0)
        if active and self.abs_agent.get() not in active: self.abs_agent.current(0)

    def add_agent(self):
        name=self.new_agent.get().strip().upper()
        if not name: messagebox.showwarning('Alta','Ingresá el nombre del agente.'); return
        if name in self.active_agents: messagebox.showwarning('Alta','Ese agente ya está activo.'); return
        if self.agent_type.get()=='ADMINISTRADOR FIJO':
            if self.admin in self.active_agents: messagebox.showwarning('Alta','Ya existe un administrador fijo activo.'); return
            self.admin=name
        else: self.sup.append(name)
        self.active_agents.add(name); self.licenses.setdefault(name,set()); self.unavailable.setdefault(name,set()); self.new_agent.set(''); self.refresh_agents(); self.status.set(f'Agente dado de alta: {name}')

    def remove_agent(self):
        sel=self.agent_tree.selection()
        if not sel: messagebox.showwarning('Baja','Seleccioná un agente.'); return
        name=self.agent_tree.item(sel[0],'values')[0]
        if name==self.admin: messagebox.showwarning('Baja','El administrador fijo debe reemplazarse antes de darlo de baja.'); return
        if not messagebox.askyesno('Confirmar baja',f'¿Dar de baja a {name}?\nNo se borrará su historial.'): return
        self.active_agents.discard(name); self.status.set(f'Agente dado de baja: {name}'); self.refresh_agents()

    def parse_month(self):
        self.year=int(self.year_var.get()); self.month=self.month_cb.current()+1
        nd=calendar.monthrange(self.year,self.month)[1]; self.holidays={int(x.strip()) for x in self.holiday_var.get().split(',') if x.strip()}
        self.holidays={d for d in self.holidays if 1<=d<=nd}; return nd
    def apply_holidays(self): self.parse_month(); self.show_calendar(); self.status.set('Feriados aplicados al calendario.')

    def show_calendar(self):
        if not hasattr(self,'cal_frame'): return
        nd=self.parse_month()
        for w in self.cal_frame.winfo_children(): w.destroy()
        agent=self.cal_agent.get() if hasattr(self,'cal_agent') else None
        if not agent: return
        mode=self.cal_mode.get(); selected=self.licenses.setdefault(agent,set()) if mode=='LICENCIA' else self.unavailable.setdefault(agent,set())
        self.calendar_agent=agent
        heads=['L','M','X','J','V','S','D']
        for c,h in enumerate(heads): ttk.Label(self.cal_frame,text=h,font=('Segoe UI',10,'bold'),anchor='center').grid(row=0,column=c,sticky='nsew',padx=3,pady=3)
        first,days=calendar.monthrange(self.year,self.month)
        for i in range(first+days):
            d=i-first+1
            if d<1: continue
            r=i//7+1; c=i%7
            state='FERIADO' if d in self.holidays else ('LICENCIA' if d in self.licenses.get(agent,set()) else ('NO DISPONIBLE' if d in self.unavailable.get(agent,set()) else 'DISPONIBLE'))
            text=str(d)
            if state=='FERIADO': bg='#FFE3A3'
            elif state=='LICENCIA': bg='#F7B7B7'
            elif state=='NO DISPONIBLE': bg='#C9D7FF'
            else: bg='#F3F5F7'
            b=tk.Button(self.cal_frame,text=text,bg=bg,activebackground=bg,relief='flat',font=('Segoe UI',11,'bold'),width=8,height=3,command=lambda x=d:self.toggle_day(x))
            b.grid(row=r,column=c,sticky='nsew',padx=4,pady=4)
        for c in range(7): self.cal_frame.columnconfigure(c,weight=1)
        self.cal_help.set(f'{agent}: modo {mode}. Rojo = licencia · Azul = no disponible · Amarillo = feriado.')

    def toggle_day(self,d):
        agent=self.cal_agent.get(); mode=self.cal_mode.get()
        if d in self.holidays: return
        target=self.licenses.setdefault(agent,set()) if mode=='LICENCIA' else self.unavailable.setdefault(agent,set())
        target.remove(d) if d in target else target.add(d)
        self.show_calendar()

    def clear_agent_month(self):
        a=self.cal_agent.get()
        if not a:return
        self.licenses[a].clear(); self.unavailable[a].clear(); self.show_calendar()

    def build_blocks(self,nd):
        blocks={}
        for a in self.all_agents():
            lic=set(d for d in self.licenses.get(a,set()) if d<=nd); no=set(d for d in self.unavailable.get(a,set()) if d<=nd); ext=set()
            for d in lic:
                if d<nd:
                    nxt=d+1
                    if (date(self.year,self.month,nxt).weekday()>=5 or nxt in self.holidays) and nxt not in lic: ext.add(nxt)
            blocks[a]={'lic':lic,'no':no,'ext':ext,'all':lic|no|ext}
        return blocks

    def pay_type(self,d,shift,aero):
        wd=date(self.year,self.month,d).weekday()
        if d in self.holidays or wd==6:return '100%'
        if wd==5 and ((aero and shift in ('13 A 19','19 A 01')) or (not aero and shift=='15 A 22')): return '100%'
        return '50%'

    def generate(self):
        try:
            nd=self.parse_month(); self.v50=float(self.v50_var.get().replace(',','.')); self.v100=float(self.v100_var.get().replace(',','.'))
            agents=self.all_agents(); blocks=self.build_blocks(nd); avail={a:max(1,nd-len(blocks[a]['all'])) for a in agents}; cron={d:{} for d in range(1,nd+1)}; zone_hist={a:{z:0 for z in range(1,7)} for a in agents}; counts={a:{k:0 for k in AERO+['SEC_50','SEC_100','TOTAL_TURNOS','H50','H100']} for a in agents}
            def score(a,key,zone=None):
                c=counts[a]; specific=c.get(key,c.get('SEC_50',0)+c.get('SEC_100',0)) if key.startswith('AERO_') else c.get(key,0); return (specific, c['TOTAL_TURNOS']/avail[a], (c['H50']+c['H100'])/avail[a], (c['H50']*self.v50+c['H100']*self.v100)/avail[a], zone_hist[a].get(zone,0) if zone else 0)
            zone_castro=1
            for d in range(1,nd+1):
                # Deliberadamente NO se consulta d-1 cuando d == 1: cada mes comienza de cero.
                active=['01 A 07','07 A 13','13 A 19','19 A 01'] if date(self.year,self.month,d).weekday()>=5 or d in self.holidays else ['01 A 07','19 A 01']
                late=set()
                if d>1:
                    late={cron[d-1].get('AERO_19_01')}|{a for k,a in cron[d-1].items() if k.startswith('ZONA_')}
                for shift in active:
                    key='AERO_'+shift.replace(' A ','_'); cand=[a for a in self.sup if a in agents and d not in blocks[a]['all'] and a not in cron[d].values()]
                    if d>1: cand=[a for a in cand if a not in {cron[d-1].get(k) for k in AERO}]
                    if shift=='01 A 07':
                        q=[a for a in cand if a not in late]
                        if q:cand=q
                    if cand: cron[d][key]=min(cand,key=lambda a:score(a,key))
                    if cand: counts[cron[d][key]][key]+=1; counts[cron[d][key]]['TOTAL_TURNOS']+=1; counts[cron[d][key]]['H100' if self.pay_type(d,shift,True)=='100%' else 'H50']+=6
                if self.admin in agents and d not in blocks[self.admin]['all']:
                    cron[d][f'ZONA_{zone_castro}']=self.admin; zone_hist[self.admin][zone_castro]+=1; counts[self.admin]['TOTAL_TURNOS']+=1; counts[self.admin]['H100' if self.pay_type(d,'15 A 22' if date(self.year,self.month,d).weekday()>=5 or d in self.holidays else '19 A 02',False)=='100%' else 'H50']+=7; zone_castro=zone_castro+1 if zone_castro<6 else 1
                occupied=zone_castro-1 if zone_castro>1 else 6; zones=[z for z in range(1,7) if z!=occupied] if self.admin in agents and d not in blocks[self.admin]['all'] else list(range(1,7))
                for z in zones:
                    shift='15 A 22' if date(self.year,self.month,d).weekday()>=5 or d in self.holidays else '19 A 02'; key='SEC_100' if self.pay_type(d,shift,False)=='100%' else 'SEC_50'; cand=[a for a in agents if d not in blocks[a]['all'] and a not in cron[d].values()]
                    if cand:
                        chosen=min(cand,key=lambda a:score(a,key,z)); cron[d][f'ZONA_{z}']=chosen; zone_hist[chosen][z]+=1; counts[chosen][key]+=1; counts[chosen]['TOTAL_TURNOS']+=1; counts[chosen]['H100' if key=='SEC_100' else 'H50']+=7
            self.result={'year':self.year,'month':self.month,'days':nd,'cron':cron,'blocks':blocks,'avail':avail,'counts':counts,'zone_hist':zone_hist}; self.replacement_log=[]; self.refresh_schedule(); self.status.set(f'Cronograma generado: {MONTHS[self.month-1]} {self.year}.'); self.nb.select(self.tab_schedule)
        except Exception as e: messagebox.showerror('Error',str(e))

    def candidate_score(self,a,key,zone):
        r=self.result;c=r['counts'][a]; av=r['avail'][a]; specific=c.get(key,0); return (specific,c['TOTAL_TURNOS']/av,(c['H50']+c['H100'])/av,(c['H50']*self.v50+c['H100']*self.v100)/av,r['zone_hist'][a].get(zone,0) if zone else 0)

    def search_replacement(self):
        if not self.result: messagebox.showwarning('Ausencias','Primero generá el cronograma.'); return
        d=self.abs_day.get(); absent=self.abs_agent.get(); requested=self.abs_key.get(); cron=self.result['cron']
        if d not in cron: return
        targets=[k for k,v in cron[d].items() if v==absent] if requested=='TODOS LOS TURNOS' else [requested]
        for x in self.abs_tree.get_children(): self.abs_tree.delete(x)
        self.pending=[]
        for key in targets:
            if cron[d].get(key)!=absent: continue
            aero=key.startswith('AERO_'); pool=self.sup if aero else self.all_agents(); zone=int(key.split('_')[1]) if key.startswith('ZONA_') else None
            cand=[]
            for a in pool:
                if a==absent or a not in self.active_agents or d in self.result['blocks'][a]['all'] or a in cron[d].values(): continue
                if d>1 and key=='AERO_01_07' and (cron[d-1].get('AERO_19_01')==a or any(v==a for k,v in cron[d-1].items() if k.startswith('ZONA_'))): continue
                if d>1 and key.startswith('AERO_') and any(cron[d-1].get(k)==a for k in AERO): continue
                cand.append(a)
            cand.sort(key=lambda a:self.candidate_score(a,'SEC_100' if key.startswith('ZONA_') and self.pay_type(d,'15 A 22' if d in self.holidays or date(self.year,self.month,d).weekday()>=5 else '19 A 02',False)=='100%' else ('SEC_50' if key.startswith('ZONA_') else key),zone))
            if cand:
                best=cand[0]; load=str(self.candidate_score(best,key if key.startswith('AERO_') else ('SEC_100' if self.pay_type(d,'15 A 22' if d in self.holidays or date(self.year,self.month,d).weekday()>=5 else '19 A 02',False)=='100%' else 'SEC_50'),zone))
                item=self.abs_tree.insert('', 'end', values=(d,key,self.turn_from_key(d,key),absent,best,load,'PENDIENTE')); self.pending.append({'item':item,'day':d,'key':key,'old':absent,'new':best})
                for alt in cand[1:5]: self.abs_tree.insert('', 'end', values=(d,key,self.turn_from_key(d,key),absent,alt,str(self.candidate_score(alt,key,zone)),'ALTERNATIVA'))
        self.abs_msg.set(f'Se encontraron {len(self.pending)} reemplazo(s) recomendado(s). Seleccioná uno y confirmá.')

    def turn_from_key(self,d,key):
        if key.startswith('AERO_'): return key.replace('AERO_','').replace('_',' A ')
        return '15 A 22' if date(self.year,self.month,d).weekday()>=5 or d in self.holidays else '19 A 02'

    def confirm_replacement(self):
        sel=self.abs_tree.selection()
        if not sel: messagebox.showwarning('Reemplazo','Seleccioná una recomendación.'); return
        vals=self.abs_tree.item(sel[0],'values')
        match=next((x for x in self.pending if x['item']==sel[0]),None)
        if not match: messagebox.showwarning('Reemplazo','Seleccioná la fila marcada como PENDIENTE.'); return
        if not messagebox.askyesno('Confirmar',f"{vals[3]} será reemplazado por {vals[4]} en {vals[1]} del día {vals[0]}.\n\nNo se modificará ningún otro puesto. ¿Confirmar?"): return
        self.result['cron'][match['day']][match['key']]=match['new']; self.replacement_log.append({'Día':match['day'],'Puesto':match['key'],'Turno':self.turn_from_key(match['day'],match['key']),'Original':match['old'],'Reemplazante':match['new'],'Motivo':'Ausencia','Fecha de registro':date.today().isoformat()}); self.recalculate_counts(); self.refresh_schedule(); self.search_replacement(); self.status.set(f"Reemplazo confirmado: {match['old']} → {match['new']}.")

    def recalculate_counts(self):
        if not self.result:return
        agents=self.all_agents(); c={a:{k:0 for k in AERO+['SEC_50','SEC_100','TOTAL_TURNOS','H50','H100']} for a in agents}; zh={a:{z:0 for z in range(1,7)} for a in agents}
        for d,ass in self.result['cron'].items():
            for key,a in ass.items():
                if a not in c: continue
                if key.startswith('AERO_'): sh=self.turn_from_key(d,key); p=self.pay_type(d,sh,True); c[a][key]+=1; c[a]['H100' if p=='100%' else 'H50']+=6
                else: sh=self.turn_from_key(d,key); p=self.pay_type(d,sh,False); c[a]['SEC_100' if p=='100%' else 'SEC_50']+=1; c[a]['H100' if p=='100%' else 'H50']+=7; zh[a][int(key.split('_')[1])]+=1
                c[a]['TOTAL_TURNOS']+=1
        self.result['counts']=c; self.result['zone_hist']=zh

    def refresh_schedule(self):
        if not self.result or not hasattr(self,'schedule_tree'):return
        for x in self.schedule_tree.get_children(): self.schedule_tree.delete(x)
        for d in sorted(self.result['cron']):
            for key,a in sorted(self.result['cron'][d].items()):
                aero=key.startswith('AERO_'); sec='AEROPUERTO' if aero else 'ZONA SECUNDARIA'; zone='AEROPUERTO' if aero else key.replace('ZONA_','ZONA '); sh=self.turn_from_key(d,key); p=self.pay_type(d,sh,aero); self.schedule_tree.insert('', 'end', values=(d,a,sec,zone,sh,p))

    def show_history(self):
        win=tk.Toplevel(self.root); win.title('Historial de reemplazos'); win.geometry('900x400'); tr=ttk.Treeview(win,columns=('Día','Puesto','Turno','Original','Reemplazante','Motivo','Fecha'),show='headings')
        for c in ('Día','Puesto','Turno','Original','Reemplazante','Motivo','Fecha'):tr.heading(c,text=c);tr.column(c,width=120,anchor='center')
        tr.pack(fill='both',expand=True,padx=10,pady=10)
        for r in self.replacement_log:tr.insert('', 'end', values=tuple(r.values()))

    def export_excel(self):
        if not self.result: messagebox.showwarning('Excel','Primero generá el cronograma.'); return
        path=filedialog.asksaveasfilename(defaultextension='.xlsx',filetypes=[('Excel','*.xlsx')],initialfile=f'GiroDeSupervisores_{self.month}_{self.year}.xlsx')
        if not path:return
        r=self.result; wb=Workbook(); ws=wb.active; ws.title='Cronograma'; headers=['DÍA','AGENTE','SECCIÓN','ZONA','TURNO','PAGO']; ws.append(headers)
        for d in sorted(r['cron']):
            for key,a in sorted(r['cron'][d].items()):
                aero=key.startswith('AERO_'); sh=self.turn_from_key(d,key); ws.append([d,a,'AEROPUERTO' if aero else 'ZONA SECUNDARIA','AEROPUERTO' if aero else key.replace('ZONA_','ZONA '),sh,self.pay_type(d,sh,aero)])
        self.style_sheet(ws)
        w2=wb.create_sheet('Control de Equidad'); w2.append(['Agente','Días disponibles','AERO 01-07','AERO 07-13','AERO 13-19','AERO 19-01','SEC 50','SEC 100','Total turnos','Horas 50','Horas 100','Horas totales','Valor 50','Valor 100','Valor total'])
        for a in self.all_agents():
            c=r['counts'][a]; w2.append([a,r['avail'][a],c['AERO_01_07'],c['AERO_07_13'],c['AERO_13_19'],c['AERO_19_01'],c['SEC_50'],c['SEC_100'],c['TOTAL_TURNOS'],c['H50'],c['H100'],c['H50']+c['H100'],c['H50']*self.v50,c['H100']*self.v100,c['H50']*self.v50+c['H100']*self.v100])
        self.style_sheet(w2)
        w3=wb.create_sheet('Licencias y No Disponibles'); w3.append(['Agente','Licencias','No disponibles','Extensión finde/feriado'])
        for a in self.all_agents():w3.append([a,','.join(map(str,sorted(r['blocks'][a]['lic']))),','.join(map(str,sorted(r['blocks'][a]['no']))),','.join(map(str,sorted(r['blocks'][a]['ext'])))])
        self.style_sheet(w3)
        w4=wb.create_sheet('Historial Reemplazos');
        if self.replacement_log:
            w4.append(list(self.replacement_log[0].keys())); [w4.append(list(x.values())) for x in self.replacement_log]
        else:w4.append(['Sin reemplazos registrados'])
        self.style_sheet(w4); wb.save(path); messagebox.showinfo('Excel','Archivo generado correctamente.')

    @staticmethod
    def style_sheet(ws):
        fill=PatternFill('solid',fgColor='172B4D'); font=Font(name='Segoe UI',bold=True,color='FFFFFF'); border=Border(*( [Side(style='thin',color='D5D9E0')]*4 ))
        for cell in ws[1]: cell.fill=fill; cell.font=font; cell.alignment=Alignment(horizontal='center')
        for row in ws.iter_rows():
            for c in row: c.border=border
        for col in ws.columns:
            letter=col[0].column_letter; ws.column_dimensions[letter].width=min(42,max(12,max(len(str(c.value or '')) for c in col)+3))

if __name__=='__main__':
    root=tk.Tk(); GiroApp(root); root.mainloop()
