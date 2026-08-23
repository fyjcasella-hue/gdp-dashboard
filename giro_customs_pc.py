import calendar
import random
from datetime import date
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

MONTHS={1:'ENERO',2:'FEBRERO',3:'MARZO',4:'ABRIL',5:'MAYO',6:'JUNIO',7:'JULIO',8:'AGOSTO',9:'SEPTIEMBRE',10:'OCTUBRE',11:'NOVIEMBRE',12:'DICIEMBRE'}
SUPERVISORES=['ACCIETTO L.','ECHAVARRIA A.','AMAYA S.','MOCAYAR L.','MINGONE G.','GARAYZABAL D.','PEPA E.','DEVOTTO M.','RODRIGUEZ J.','MARTINEZ PAZ S.','DOMINGUEZ V.','BUSTOS FIERRO F.','JANISZEWSKI J.','MERLO C.','URZAGASTI F.']
ADMIN='CASTRO D.'
AGENTS=SUPERVISORES+[ADMIN]
DEFAULT_LIC={'DOMINGUEZ V.':[2,3,4,5,6,7,8,9],'PEPA E.':[15,16,17,18,19,20,21,22,23],'RODRIGUEZ J.':list(range(1,15)),'URZAGASTI F.':[4,5],'GARAYZABAL D.':[13,14,15,16,17,18,19,20]}
DEFAULT_NO={'MERLO C.':[8,9,15,16,17,29,30],'DOMINGUEZ V.':[24,29],'JANISZEWSKI J.':[15,16,17],'URZAGASTI F.':[15,16,17]}
AERO=['AERO_01_07','AERO_07_13','AERO_13_19','AERO_19_01']

class GiroCustomsApp:
    def __init__(self,root):
        self.root=root; self.root.title('GIRO CUSTOMS — Generador y Reemplazos'); self.root.geometry('1180x760'); self.root.minsize(1000,680)
        self.result=None; self.replacements=[]; self.build_ui()

    def build_ui(self):
        s=ttk.Style()
        try:s.theme_use('clam')
        except tk.TclError:pass
        head=tk.Frame(self.root,bg='#1F4E79',height=82);head.pack(fill='x')
        tk.Label(head,text='GIRO CUSTOMS',bg='#1F4E79',fg='white',font=('Segoe UI',22,'bold')).pack(anchor='w',padx=28,pady=(12,0))
        tk.Label(head,text='Cronograma · Equidad · Ausencias · Reemplazo local',bg='#1F4E79',fg='#D9EAF7',font=('Segoe UI',10)).pack(anchor='w',padx=30)
        nb=ttk.Notebook(self.root);nb.pack(fill='both',expand=True,padx=18,pady=14)
        self.tab_config=ttk.Frame(nb,padding=18);self.tab_abs=ttk.Frame(nb,padding=18);self.tab_preview=ttk.Frame(nb,padding=12)
        nb.add(self.tab_config,text='Configuración y generación');nb.add(self.tab_abs,text='Ausencias y reemplazos');nb.add(self.tab_preview,text='Vista del cronograma')
        self.nb=nb;self.build_config();self.build_absences();self.build_preview()

    def build_config(self):
        f=ttk.LabelFrame(self.tab_config,text='Período y valores',padding=14);f.pack(fill='x')
        ttk.Label(f,text='Año').grid(row=0,column=0,padx=6,pady=6,sticky='w');self.year=ttk.Spinbox(f,from_=2020,to=2100,width=9);self.year.set('2026');self.year.grid(row=0,column=1)
        ttk.Label(f,text='Mes').grid(row=0,column=2,padx=6);self.month=ttk.Combobox(f,values=[f'{i:02d} - {MONTHS[i]}' for i in range(1,13)],state='readonly',width=18);self.month.current(7);self.month.grid(row=0,column=3)
        ttk.Label(f,text='Feriados').grid(row=0,column=4,padx=6);self.holidays=ttk.Entry(f,width=16);self.holidays.insert(0,'17');self.holidays.grid(row=0,column=5)
        ttk.Label(f,text='Hora 50%').grid(row=1,column=0,padx=6,pady=6,sticky='w');self.v50=ttk.Entry(f,width=12);self.v50.insert(0,'17731.5');self.v50.grid(row=1,column=1)
        ttk.Label(f,text='Hora 100%').grid(row=1,column=2,padx=6);self.v100=ttk.Entry(f,width=12);self.v100.insert(0,'23642');self.v100.grid(row=1,column=3)
        ttk.Label(f,text='Semilla opcional').grid(row=1,column=4,padx=6);self.seed=ttk.Entry(f,width=16);self.seed.grid(row=1,column=5)
        p=ttk.LabelFrame(self.tab_config,text='Licencias y días no trabajables',padding=14);p.pack(fill='both',expand=True,pady=14)
        l=ttk.Frame(p);l.pack(side='left',fill='both',expand=True,padx=(0,8));ttk.Label(l,text='LICENCIAS — NOMBRE: días').pack(anchor='w');self.lic=tk.Text(l,font=('Consolas',9));self.lic.pack(fill='both',expand=True);self.lic.insert('1.0','\n'.join(f'{a}: {",".join(map(str,d))}' for a,d in DEFAULT_LIC.items()))
        r=ttk.Frame(p);r.pack(side='left',fill='both',expand=True,padx=(8,0));ttk.Label(r,text='NO TRABAJABLES — NOMBRE: días').pack(anchor='w');self.no=tk.Text(r,font=('Consolas',9));self.no.pack(fill='both',expand=True);self.no.insert('1.0','\n'.join(f'{a}: {",".join(map(str,d))}' for a,d in DEFAULT_NO.items()))
        bar=ttk.Frame(self.tab_config);bar.pack(fill='x');self.status=tk.StringVar(value='Listo.');ttk.Label(bar,textvariable=self.status).pack(side='left');ttk.Button(bar,text='GENERAR CRONOGRAMA',command=self.generate).pack(side='right',padx=4);ttk.Button(bar,text='EXPORTAR EXCEL',command=self.export_excel).pack(side='right',padx=4)

    def build_absences(self):
        info=ttk.LabelFrame(self.tab_abs,text='Reemplazo automático de una ausencia',padding=14);info.pack(fill='x')
        ttk.Label(info,text='El cronograma ya generado no se regenera. Solo se modifica el puesto afectado.').pack(anchor='w')
        row=ttk.Frame(info);row.pack(fill='x',pady=12)
        ttk.Label(row,text='Día').pack(side='left');self.abs_day=ttk.Spinbox(row,from_=1,to=31,width=6);self.abs_day.set('1');self.abs_day.pack(side='left',padx=6)
        ttk.Label(row,text='Agente ausente').pack(side='left',padx=(18,4));self.abs_agent=ttk.Combobox(row,values=AGENTS,state='readonly',width=25);self.abs_agent.current(0);self.abs_agent.pack(side='left')
        ttk.Label(row,text='Puesto/turno').pack(side='left',padx=(18,4));self.abs_key=ttk.Combobox(row,state='readonly',width=22);self.abs_key['values']=('TODOS LOS TURNOS',*AERO,*[f'ZONA_{i}' for i in range(1,7)]);self.abs_key.current(0);self.abs_key.pack(side='left')
        ttk.Button(row,text='BUSCAR REEMPLAZO',command=self.search_replacement).pack(side='left',padx=12)
        ttk.Button(row,text='CONFIRMAR REEMPLAZO',command=self.confirm_replacement).pack(side='left')
        cols=('Día','Puesto','Turno','Original','Candidato','Puntuación','Estado');self.candidates=ttk.Treeview(self.tab_abs,columns=cols,show='headings',height=14)
        for c,w in zip(cols,(55,150,100,160,170,240,120)):self.candidates.heading(c,text=c);self.candidates.column(c,width=w,anchor='center')
        self.candidates.pack(fill='both',expand=True,pady=12)
        ttk.Label(self.tab_abs,text='Historial de reemplazos confirmados').pack(anchor='w');hcols=('Día','Puesto','Original','Reemplazo','Turno','Pago');self.history=ttk.Treeview(self.tab_abs,columns=hcols,show='headings',height=6)
        for c,w in zip(hcols,(55,150,170,170,100,80)):self.history.heading(c,text=c);self.history.column(c,width=w,anchor='center')
        self.history.pack(fill='x')
        self.pending=[]

    def build_preview(self):
        self.preview_tree=ttk.Treeview(self.tab_preview,columns=('Día','Agente','Sección','Zona','Turno','Pago'),show='headings')
        for c,w in zip(('Día','Agente','Sección','Zona','Turno','Pago'),(60,180,150,150,110,80)):self.preview_tree.heading(c,text=c);self.preview_tree.column(c,width=w,anchor='center')
        self.preview_tree.pack(fill='both',expand=True)

    def parse_days(self,text):
        out={}
        for line in text.splitlines():
            if ':' not in line or not line.strip():continue
            n,v=line.split(':',1);n=n.strip()
            try:out[n]=sorted({int(x.strip()) for x in v.split(',') if x.strip()})
            except ValueError:raise ValueError(f'Días inválidos en {n}.')
        return out

    def config(self):
        y=int(self.year.get());m=self.month.current()+1;days=calendar.monthrange(y,m)[1]
        hs={int(x.strip()) for x in self.holidays.get().split(',') if x.strip()}
        if any(d<1 or d>days for d in hs):raise ValueError('Hay un feriado fuera del mes seleccionado.')
        v50=float(self.v50.get().replace(',','.'));v100=float(self.v100.get().replace(',','.'));seed=self.seed.get().strip();random.seed(int(seed) if seed else None)
        return y,m,days,hs,v50,v100,self.parse_days(self.lic.get('1.0','end')),self.parse_days(self.no.get('1.0','end'))

    @staticmethod
    def ranges(vals):
        if not vals:return 'Ninguno'
        d=sorted(set(vals));out=[];a=b=d[0]
        for x in d[1:]:
            if x==b+1:b=x
            else:out.append(str(a) if a==b else f'{a} al {b}');a=b=x
        out.append(str(a) if a==b else f'{a} al {b}');return ', '.join(out)

    def generate(self):
        try:
            y,m,nd,hol,v50,v100,lic,no=self.config();week=lambda d:date(y,m,d).weekday()>=5 or d in hol
            blocked={}
            for a in AGENTS:
                l=set(lic.get(a,[]));n=set(no.get(a,[]));e={d+1 for d in l if d<nd and week(d+1) and d+1 not in l};blocked[a]={'todos':l|e|n,'lic':l,'ext':e,'no':n}
            avail={a:max(1,nd-len(blocked[a]['todos'])) for a in AGENTS};cron={d:{} for d in range(1,nd+1)};zone_hist={a:{z:0 for z in range(1,7)} for a in AGENTS};zone_castro=1
            def pay(d,t,aero):
                wd=date(y,m,d).weekday()
                if d in hol or wd==6:return '100%'
                if wd==5 and ((aero and t in ('13 A 19','19 A 01')) or (not aero and t=='15 A 22')):return '100%'
                return '50%'
            def counts():
                c={a:{k:0 for k in AERO+['SEC_50','SEC_100','TOTAL_TURNOS','HORAS_50','HORAS_100']} for a in AGENTS}
                for d,ass in cron.items():
                    for k,a in ass.items():
                        if k.startswith('AERO_'):c[a][k]+=1;c[a]['TOTAL_TURNOS']+=1;c[a]['HORAS_100' if pay(d,k.replace('AERO_','').replace('_',' A '),True)=='100%' else 'HORAS_50']+=6
                        else:c[a]['SEC_100' if pay(d,'15 A 22' if week(d) else '19 A 02',False)=='100%' else 'SEC_50']+=1;c[a]['TOTAL_TURNOS']+=1;c[a]['HORAS_100' if pay(d,'15 A 22' if week(d) else '19 A 02',False)=='100%' else 'HORAS_50']+=7;zone_hist[a][int(k.split('_')[1])]+=1
                for a in AGENTS:c[a]['INGRESOS']=c[a]['HORAS_50']*v50+c[a]['HORAS_100']*v100
                return c
            for d in range(1,nd+1):
                late=set()
                if d>1:late|={cron[d-1].get('AERO_19_01')};late|={a for k,a in cron[d-1].items() if k.startswith('ZONA_')}
                active=('01 A 07','07 A 13','13 A 19','19 A 01') if week(d) else ('01 A 07','19 A 01')
                for t in active:
                    key='AERO_'+t.replace(' A ','_');c=counts();cand=[a for a in SUPERVISORES if d not in blocked[a]['todos'] and a not in cron[d].values()]
                    if d>1:cand=[a for a in cand if a not in {cron[d-1].get(k) for k in AERO}]
                    if t=='01 A 07':
                        preferred=[a for a in cand if a not in late]
                        if preferred:cand=preferred
                    cand.sort(key=lambda a:(c[a][key],c[a]['TOTAL_TURNOS']/avail[a],c[a]['HORAS_100']+c[a]['HORAS_50'],c[a]['INGRESOS']/avail[a]))
                    if cand:cron[d][key]=cand[0]
                if d not in blocked[ADMIN]['todos']:
                    t='15 A 22' if week(d) else '19 A 02';cron[d][f'ZONA_{zone_castro}']=ADMIN;zone_castro=zone_castro+1 if zone_castro<6 else 1
                occupied=zone_castro-1 if zone_castro>1 else 6
                zones=[1,2,3,4,5,6] if d in blocked[ADMIN]['todos'] else [z for z in range(1,7) if z!=occupied]
                for z in zones:
                    c=counts();cand=[a for a in AGENTS if d not in blocked[a]['todos'] and a not in cron[d].values()];cand.sort(key=lambda a:(c[a]['INGRESOS']/avail[a],c[a]['TOTAL_TURNOS']/avail[a],zone_hist[a][z]))
                    if cand:cron[d][f'ZONA_{z}']=cand[0]
            self.result=self.rebuild_result(y,m,nd,hol,v50,v100,blocked,avail,cron,zone_hist);self.replacements=[];self.refresh_preview();self.status.set(f'Cronograma generado: {MONTHS[m]} {y}.');messagebox.showinfo('GIRO CUSTOMS','Cronograma generado correctamente.')
        except Exception as e:messagebox.showerror('Error',str(e))

    def rebuild_result(self,y,m,nd,hol,v50,v100,blocked,avail,cron,zone_hist):
        def weekend(d):return date(y,m,d).weekday()>=5 or d in hol
        def pay(d,t,aero):
            wd=date(y,m,d).weekday()
            if d in hol or wd==6:return '100%'
            if wd==5 and ((aero and t in ('13 A 19','19 A 01')) or (not aero and t=='15 A 22')):return '100%'
            return '50%'
        rows=[]
        for d,ass in cron.items():
            for k,a in ass.items():
                if k.startswith('AERO_'):sec='AEROPUERTO';turn=k.replace('AERO_','').replace('_',' A ');zone='AEROPUERTO';p=pay(d,turn,True)
                else:sec='ZONA SECUNDARIA';turn='15 A 22' if weekend(d) else '19 A 02';zone=k.replace('ZONA_','ZONA ');p=pay(d,turn,False)
                rows.append({'DÍA':d,'AGENTE':a,'SECCIÓN':sec,'ZONA':zone,'TURNO':turn,'PAGO':p})
        c={a:{k:0 for k in AERO+['SEC_50','SEC_100','TOTAL_TURNOS','CANT_HORAS_50','CANT_HORAS_100']} for a in AGENTS}
        for r in rows:
            a=r['AGENTE'];hours=6 if r['SECCIÓN']=='AEROPUERTO' else 7
            if r['SECCIÓN']=='AEROPUERTO':c[a]['AERO_'+r['TURNO'].replace(' A ','_')]+=1
            else:c[a]['SEC_100' if r['PAGO']=='100%' else 'SEC_50']+=1
            c[a]['TOTAL_TURNOS']+=1;c[a]['CANT_HORAS_100' if r['PAGO']=='100%' else 'CANT_HORAS_50']+=hours
        control=[]
        for a in AGENTS:
            x=c[a];h50=x['CANT_HORAS_50'];h100=x['CANT_HORAS_100'];v50t=h50*v50;v100t=h100*v100;total=v50t+v100t
            control.append({'Agente':a,'Dias_Disponibles':avail[a],**{k:x[k] for k in AERO+['SEC_50','SEC_100','TOTAL_TURNOS']},'CANT_HORAS_50':h50,'CANT_HORAS_100':h100,'CANT_HORAS_TOTALES':h50+h100,'VALOR_TOTAL_50':v50t,'VALOR_TOTAL_100':v100t,'VALOR_TOTAL_AGENTE':total,'Sueldo_Proporcional_Por_Dia':round(total/avail[a],2)})
        lic=[]
        for a in AGENTS:
            b=blocked[a];lic.append({'Agente':a,'Días Licencia':len(b['lic']),'Fechas Licencia':self.ranges(b['lic']),'Días Extensión (Finde/Feriado)':len(b['ext']),'Fechas Extensión Finde/Feriado':self.ranges(b['ext']),'Días No Trabajables':len(b['no']),'Fechas Días No Trabajables':self.ranges(b['no']),'Total Días No Disponibles':len(b['todos'])})
        return {'year':y,'month':m,'days':nd,'holidays':hol,'v50':v50,'v100':v100,'blocked':blocked,'avail':avail,'cron':cron,'zone_hist':zone_hist,'df_cron':pd.DataFrame(rows),'df_control':pd.DataFrame(control),'df_lic':pd.DataFrame(lic)}

    def refresh_preview(self):
        for i in self.preview_tree.get_children():self.preview_tree.delete(i)
        if not self.result:return
        for r in self.result['df_cron'].itertuples(index=False):self.preview_tree.insert('', 'end',values=tuple(r))

    def replacement_candidates(self,day,key,absent):
        r=self.result;y=r['year'];m=r['month'];cron=r['cron'];blocked=r['blocked'];avail=r['avail'];week=date(y,m,day).weekday()>=5 or day in r['holidays']
        if key.startswith('AERO_'):pool=SUPERVISORES;turn=key.replace('AERO_','').replace('_',' A ');aero=True
        else:pool=AGENTS;turn='15 A 22' if week else '19 A 02';aero=False
        def score(a):
            df=r['df_control'];row=df[df['Agente']==a].iloc[0];specific=row[key] if key in row.index else row['SEC_100' if self.payment(day,turn,aero)=='100%' else 'SEC_50'];zone=0
            if key.startswith('ZONA_'):zone=r['zone_hist'][a][int(key.split('_')[1])]
            return (specific,row['TOTAL_TURNOS']/avail[a],row['CANT_HORAS_TOTALES']/avail[a],row['VALOR_TOTAL_AGENTE']/avail[a],zone)
        cand=[]
        for a in pool:
            if a==absent or day in blocked[a]['todos'] or a in cron[day].values():continue
            if day>1:
                prev=cron[day-1]
                if key=='AERO_01_07' and (prev.get('AERO_19_01')==a or any(v==a for k,v in prev.items() if k.startswith('ZONA_'))):continue
                if key.startswith('AERO_') and any(prev.get(k)==a for k in AERO):continue
                if key.startswith('ZONA_') and any(v==a for k,v in prev.items() if k.startswith('ZONA_')):continue
            cand.append((score(a),a))
        return sorted(cand)

    def payment(self,d,t,aero):
        y=self.result['year'];m=self.result['month'];h=self.result['holidays'];wd=date(y,m,d).weekday()
        if d in h or wd==6:return '100%'
        if wd==5 and ((aero and t in ('13 A 19','19 A 01')) or (not aero and t=='15 A 22')):return '100%'
        return '50%'

    def search_replacement(self):
        if not self.result:return messagebox.showwarning('Sin cronograma','Primero genere el cronograma.')
        try:day=int(self.abs_day.get())
        except ValueError:return messagebox.showerror('Dato inválido','El día debe ser numérico.')
        if day<1 or day>self.result['days']:return messagebox.showerror('Dato inválido','El día no pertenece al mes.')
        absent=self.abs_agent.get();sel=self.abs_key.get();keys=[k for k,a in self.result['cron'][day].items() if a==absent] if sel=='TODOS LOS TURNOS' else [sel]
        for i in self.candidates.get_children():self.candidates.delete(i)
        self.pending=[]
        if not keys:return messagebox.showwarning('Sin asignación',f'{absent} no tiene el puesto seleccionado el día {day}.')
        for key in keys:
            if self.result['cron'][day].get(key)!=absent:continue
            ranked=self.replacement_candidates(day,key,absent)
            for rank,(score,a) in enumerate(ranked[:8]):
                turn=key.replace('AERO_','').replace('_',' A ') if key.startswith('AERO_') else ('15 A 22' if date(self.result['year'],self.result['month'],day).weekday()>=5 or day in self.result['holidays'] else '19 A 02')
                iid=self.candidates.insert('', 'end',values=(day,key,turn,absent,a,str(score),'MEJOR' if rank==0 else 'ALTERNATIVA'))
                if rank==0:self.pending.append((iid,day,key,absent,a,score))
        if not self.pending:messagebox.showwarning('Sin reemplazo','No existe un candidato compatible para ese puesto.')

    def confirm_replacement(self):
        if not self.pending:return messagebox.showwarning('Sin propuesta','Primero pulse BUSCAR REEMPLAZO.')
        changes=[]
        for _,day,key,absent,a,score in self.pending:
            if self.result['cron'][day].get(key)!=absent:continue
            turn=key.replace('AERO_','').replace('_',' A ') if key.startswith('AERO_') else ('15 A 22' if date(self.result['year'],self.result['month'],day).weekday()>=5 or day in self.result['holidays'] else '19 A 02')
            pay=self.payment(day,turn,key.startswith('AERO_'));self.result['cron'][day][key]=a;changes.append((day,key,absent,a,turn,pay))
        if not changes:return messagebox.showwarning('No aplicado','La asignación cambió antes de confirmar.')
        self.rebuild_after_change();
        for day,key,old,new,turn,pay in changes:self.history.insert('', 'end',values=(day,key,old,new,turn,pay));self.replacements.append({'DÍA':day,'PUESTO':key,'ORIGINAL':old,'REEMPLAZO':new,'TURNO':turn,'PAGO':pay})
        self.pending=[];self.refresh_preview();self.status.set(f'{len(changes)} reemplazo(s) confirmado(s).');messagebox.showinfo('Reemplazo confirmado','Se modificó únicamente el/los puesto(s) afectado(s).')

    def rebuild_after_change(self):
        r=self.result;self.result=self.rebuild_result(r['year'],r['month'],r['days'],r['holidays'],r['v50'],r['v100'],r['blocked'],r['avail'],r['cron'],r['zone_hist'])

    def export_excel(self):
        if not self.result:return messagebox.showwarning('Sin datos','Primero genere el cronograma.')
        r=self.result;default=f'Giro_Customs_{MONTHS[r["month"]]}_{r["year"]}.xlsx';path=filedialog.asksaveasfilename(defaultextension='.xlsx',initialfile=default,filetypes=[('Excel','*.xlsx')])
        if not path:return
        try:
            wb=openpyxl.Workbook();ws=wb.active;ws.title='Cronograma Día a Día';mat=wb.create_sheet('Cronograma Consolidado');ctrl=wb.create_sheet('Control de Equidad');lic=wb.create_sheet('Consolidado de Licencias')
            blue=PatternFill('solid',fgColor='1F4E79');light=PatternFill('solid',fgColor='D9E1F2');green=PatternFill('solid',fgColor='E2EFDA');yellow=PatternFill('solid',fgColor='FEF2CB');white=PatternFill('solid',fgColor='FFFFFF');zebra=PatternFill('solid',fgColor='F9FAFB');border=Border(*(Side('thin',color='BDC3C7'),)*4);hf=Font(name='Segoe UI',size=10,bold=True,color='FFFFFF');df=Font(name='Segoe UI',size=9)
            ws['A1']=f'CRONOGRAMA DE ASIGNACIONES - {MONTHS[r["month"]]} {r["year"]}';ws['A1'].font=Font(name='Segoe UI',size=14,bold=True,color='1F4E79')
            self.write_df(ws,r['df_cron'],blue,hf,df,border)
            mat.merge_cells('B3:C3');mat['B3']='DIA - HORARIO A CUBRIR';mat.merge_cells('D3:G3');mat['D3']='SECCIÓN AEROPUERTO';mat.merge_cells('H3:M3');mat['H3']='ZONA SECUNDARIA'
            heads=['01 A 07','07 A 13','13 A 19','19 A 01','ZONA 1','ZONA 2','ZONA 3','ZONA 4','ZONA 5','ZONA 6']
            for i,h in enumerate(heads,4):mat.cell(4,i,h)
            days=['LUNES','MARTES','MIÉRCOLES','JUEVES','VIERNES','SÁBADO','DOMINGO']
            mapping=AERO+[f'ZONA_{i}' for i in range(1,7)]
            for d in range(1,r['days']+1):
                rr=d+4;mat.cell(rr,2,days[date(r['year'],r['month'],d).weekday()]);mat.cell(rr,3,d)
                for ci,k in enumerate(mapping,4):mat.cell(rr,ci,r['cron'][d].get(k,''));mat.cell(rr,ci).fill=yellow if date(r['year'],r['month'],d).weekday()>=5 or d in r['holidays'] else (green if ci in (4,7) else (white if rr%2==0 else zebra));mat.cell(rr,ci).border=border;mat.cell(rr,ci).alignment=Alignment(horizontal='center')
            ctrl['A1']=f'MÉTRICAS Y CONTROL DE EQUIDAD FINANCIERA - {MONTHS[r["month"]]} {r["year"]}';self.write_df(ctrl,r['df_control'],blue,hf,df,border,currency=True)
            lic['A1']=f'CONSOLIDADO DE LICENCIAS Y DÍAS NO TRABAJABLES - {MONTHS[r["month"]]} {r["year"]}';self.write_df(lic,r['df_lic'],blue,hf,df,border)
            if self.replacements:
                rep=wb.create_sheet('Historial Reemplazos');self.write_df(rep,pd.DataFrame(self.replacements),blue,hf,df,border)
            for sh in wb.worksheets:
                for col in sh.columns:
                    letter=get_column_letter(col[0].column);sh.column_dimensions[letter].width=max(12,min(42,max((len(str(c.value or '')) for c in col),default=10)+3))
            wb.save(path);self.status.set(f'Excel guardado: {path}');messagebox.showinfo('Excel','Archivo exportado correctamente.')
        except Exception as e:messagebox.showerror('Error Excel',str(e))

    @staticmethod
    def write_df(ws,df,fill,hf,dfont,border,currency=False):
        for ci,t in enumerate(df.columns,1):c=ws.cell(3,ci,t);c.fill=fill;c.font=hf;c.alignment=Alignment(horizontal='center')
        for ri,row in enumerate(df.itertuples(index=False),4):
            for ci,v in enumerate(row,1):c=ws.cell(ri,ci,v);c.font=dfont;c.border=border;c.alignment=Alignment(horizontal='center' if ci>1 else 'left');c.fill=PatternFill('solid',fgColor='F9FAFB' if ri%2==0 else 'FFFFFF')
            if currency:
                for c in ws[ri][12:]:c.number_format='$#,##0.00'

if __name__=='__main__':
    root=tk.Tk();GiroCustomsApp(root);root.mainloop()
