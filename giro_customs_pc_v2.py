import calendar
import random
from datetime import date
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

MESES={1:'ENERO',2:'FEBRERO',3:'MARZO',4:'ABRIL',5:'MAYO',6:'JUNIO',7:'JULIO',8:'AGOSTO',9:'SEPTIEMBRE',10:'OCTUBRE',11:'NOVIEMBRE',12:'DICIEMBRE'}
SUPERVISORES=['ACCIETTO L.','ECHAVARRIA A.','AMAYA S.','MOCAYAR L.','MINGONE G.','GARAYZABAL D.','PEPA E.','DEVOTTO M.','RODRIGUEZ J.','MARTINEZ PAZ S.','DOMINGUEZ V.','BUSTOS FIERRO F.','JANISZEWSKI J.','MERLO C.','URZAGASTI F.']
ADMIN='CASTRO D.'
AGENTES=SUPERVISORES+[ADMIN]
LIC={'DOMINGUEZ V.':[2,3,4,5,6,7,8,9],'PEPA E.':[15,16,17,18,19,20,21,22,23],'RODRIGUEZ J.':list(range(1,15)),'URZAGASTI F.':[4,5],'GARAYZABAL D.':[13,14,15,16,17,18,19,20]}
NO_TRAB={'MERLO C.':[8,9,15,16,17,29,30],'DOMINGUEZ V.':[24,29],'JANISZEWSKI J.':[15,16,17],'URZAGASTI F.':[15,16,17]}

class GiroCustoms:
    def __init__(self,root):
        self.root=root; self.root.title('GIRO CUSTOMS — Gestión automática de turnos'); self.root.geometry('1120x780'); self.result=None
        self.ui()
    def ui(self):
        top=tk.Frame(self.root,bg='#1F4E79',height=88);top.pack(fill='x')
        tk.Label(top,text='GIRO CUSTOMS',bg='#1F4E79',fg='white',font=('Segoe UI',23,'bold')).pack(anchor='w',padx=28,pady=(12,0))
        tk.Label(top,text='Cronograma • Equidad • Reemplazos automáticos por ausencia',bg='#1F4E79',fg='#D9EAF7',font=('Segoe UI',10)).pack(anchor='w',padx=30)
        body=ttk.Frame(self.root,padding=18);body.pack(fill='both',expand=True)
        cfg=ttk.LabelFrame(body,text='Configuración',padding=12);cfg.pack(fill='x')
        self.year=self.field(cfg,'Año','2026',0,0,10); self.month=ttk.Combobox(cfg,values=[f'{i:02d} - {MESES[i]}' for i in range(1,13)],state='readonly',width=17);self.month.current(7);self.month.grid(row=0,column=3,padx=5)
        ttk.Label(cfg,text='Mes').grid(row=0,column=2,sticky='w');self.holidays=self.field(cfg,'Feriados','17',0,4,18)
        self.v50=self.field(cfg,'Valor hora 50%','17731.5',1,0,12);self.v100=self.field(cfg,'Valor hora 100%','23642',1,2,12);self.seed=self.field(cfg,'Semilla','',1,4,18)
        panes=ttk.Frame(body);panes.pack(fill='both',expand=True,pady=12)
        self.lic=self.textbox(panes,'Licencias',LIC);self.no=self.textbox(panes,'Días no trabajables',NO_TRAB);self.abs=self.textbox(panes,'AUSENCIAS EXTRAORDINARIAS — se reemplazan automáticamente',{})
        bar=ttk.Frame(body);bar.pack(fill='x')
        self.status=tk.StringVar(value='Listo.');ttk.Label(bar,textvariable=self.status).pack(side='left')
        ttk.Button(bar,text='GENERAR / REEMPLAZAR AUSENCIAS',command=self.generate).pack(side='right',padx=4)
        ttk.Button(bar,text='VISTA PREVIA',command=self.preview).pack(side='right',padx=4)
        ttk.Button(bar,text='EXPORTAR EXCEL',command=self.export).pack(side='right',padx=4)
    def field(self,parent,label,value,row,col,width):
        ttk.Label(parent,text=label).grid(row=row,column=col,sticky='w',padx=5,pady=5);e=ttk.Entry(parent,width=width);e.insert(0,value);e.grid(row=row,column=col+1,padx=5);return e
    def textbox(self,parent,title,data):
        f=ttk.LabelFrame(parent,text=title,padding=8);f.pack(side='left',fill='both',expand=True,padx=5)
        t=tk.Text(f,height=13,font=('Consolas',9));t.pack(fill='both',expand=True)
        t.insert('1.0','\n'.join(f'{a}: {",".join(map(str,d))}' for a,d in data.items()))
        ttk.Label(f,text='Formato: AGENTE: 1,2,3').pack(anchor='w',pady=(4,0));return t
    @staticmethod
    def parse(t):
        out={}
        for line in t.get('1.0','end').splitlines():
            if ':' not in line:continue
            n,v=line.split(':',1);n=n.strip()
            if not n:continue
            try:out[n]=sorted({int(x.strip()) for x in v.split(',') if x.strip()})
            except ValueError:raise ValueError(f'Días inválidos para {n}.')
        return out
    @staticmethod
    def ranges(days):
        d=sorted(set(days))
        if not d:return 'Ninguno'
        p=[];a=b=d[0]
        for x in d[1:]:
            if x==b+1:b=x
            else:p.append(str(a) if a==b else f'{a} al {b}');a=b=x
        p.append(str(a) if a==b else f'{a} al {b}');return ', '.join(p)
    def generate(self):
        try:
            Y=int(self.year.get());M=self.month.current()+1;days=calendar.monthrange(Y,M)[1];hol={int(x) for x in self.holidays.get().split(',') if x.strip()};v50=float(self.v50.get().replace(',','.'));v100=float(self.v100.get().replace(',','.'));seed=self.seed.get().strip();random.seed(int(seed) if seed else None)
            lic=self.parse(self.lic);no=self.parse(self.no);absences=self.parse(self.abs)
            unknown=set(lic)|set(no)|set(absences)-set(AGENTES)
            if unknown:raise ValueError('Agentes no reconocidos: '+', '.join(sorted(unknown)))
            blocked={}
            def special(d):return date(Y,M,d).weekday()>=5 or d in hol
            for a in AGENTES:
                l=set(lic.get(a,[]));n=set(no.get(a,[]));ab=set(absences.get(a,[]));ext=set()
                for d in l:
                    nx=d+1
                    if nx<=days and special(nx) and nx not in l:ext.add(nx)
                blocked[a]=l|ext|n|ab
            avail={a:max(1,days-len(blocked[a])) for a in AGENTES}
            cron={d:{} for d in range(1,days+1)};zones={a:{z:0 for z in range(1,7)} for a in AGENTES};counts={a:{k:0 for k in ['AERO_01_07','AERO_07_13','AERO_13_19','AERO_19_01','SEC_50','SEC_100','TOTAL_TURNOS','HORAS_50','HORAS_100','AERO_FIN_DE_SEMANA']} for a in AGENTES};zc=1
            def pay(d,t,aero):
                wd=date(Y,M,d).weekday()
                if d in hol or wd==6:return '100%'
                if wd==5 and aero and t in ('13 A 19','19 A 01'):return '100%'
                if wd==5 and not aero and t=='15 A 22':return '100%'
                return '50%'
            def score(a,z=None):
                h=counts[a]['HORAS_50']*v50+counts[a]['HORAS_100']*v100
                return (h/avail[a],counts[a]['TOTAL_TURNOS']/avail[a],zones[a][z] if z else 0)
            for d in range(1,days+1):
                weekend=special(d);late=set()
                if d>1:late.update(a for k,a in cron[d-1].items() if k=='AERO_19_01' or k.startswith('ZONA_'))
                active=['01 A 07','07 A 13','13 A 19','19 A 01'] if weekend else ['01 A 07','19 A 01']
                for t in active:
                    key='AERO_'+t.replace(' A ','_');cand=[]
                    for a in SUPERVISORES:
                        if a in cron[d].values() or d in blocked[a]:continue
                        if d>1 and a in cron[d-1].values():continue
                        if t=='01 A 07' and a in late:continue
                        cand.append(a)
                    if not cand:cand=[a for a in SUPERVISORES if a not in cron[d].values() and d not in blocked[a]]
                    if cand:
                        random.shuffle(cand);cand.sort(key=lambda a:(counts[a][key],)+score(a));a=cand[0];cron[d][key]=a;counts[a][key]+=1;counts[a]['TOTAL_TURNOS']+=1;counts[a]['HORAS_100' if pay(d,t,True)=='100%' else 'HORAS_50']+=6;counts[a]['AERO_FIN_DE_SEMANA']+=int(weekend)
                if d not in blocked[ADMIN]:
                    t='15 A 22' if weekend else '19 A 02';typ=pay(d,t,False);cron[d][f'ZONA_{zc}']=ADMIN;counts[ADMIN]['SEC_100' if typ=='100%' else 'SEC_50']+=1;counts[ADMIN]['TOTAL_TURNOS']+=1;counts[ADMIN]['HORAS_100' if typ=='100%' else 'HORAS_50']+=7;zones[ADMIN][zc]+=1;zc=zc+1 if zc<6 else 1
                occupied=zc-1 if zc>1 else 6;remaining=[z for z in range(1,7) if d in blocked[ADMIN] or z!=occupied];random.shuffle(remaining)
                for z in remaining:
                    t='15 A 22' if weekend else '19 A 02';typ=pay(d,t,False);lbl='SEC_100' if typ=='100%' else 'SEC_50';cand=[a for a in AGENTES if a not in cron[d].values() and d not in blocked[a]]
                    if cand:
                        random.shuffle(cand);cand.sort(key=lambda a:score(a,z));a=cand[0];cron[d][f'ZONA_{z}']=a;counts[a][lbl]+=1;counts[a]['TOTAL_TURNOS']+=1;counts[a]['HORAS_100' if typ=='100%' else 'HORAS_50']+=7;zones[a][z]+=1
            # Audit: every assignment on an absence day must belong to a different eligible agent.
            replacements=[]
            for a,ad in absences.items():
                for d in ad:
                    for k,x in cron[d].items():
                        if x==a:replacements.append((d,k,a))
            if replacements:raise RuntimeError('Se detectó una ausencia aún asignada; revise las restricciones.')
            rows=[]
            for d,ass in cron.items():
                for k,a in ass.items():
                    if k.startswith('AERO_'):sec='AEROPUERTO';turn=k.replace('AERO_','').replace('_',' A ');zone='AEROPUERTO';p=pay(d,turn,True)
                    else:sec='ZONA SECUNDARIA';turn='15 A 22' if special(d) else '19 A 02';zone=k.replace('ZONA_','ZONA ');p=pay(d,turn,False)
                    rows.append({'DÍA':d,'AGENTE':a,'SECCIÓN':sec,'ZONA':zone,'TURNO':turn,'PAGO':p})
            df=pd.DataFrame(rows).sort_values(['DÍA','SECCIÓN','TURNO']);control=[]
            for a in AGENTES:
                h50=counts[a]['HORAS_50'];h100=counts[a]['HORAS_100'];total=h50*v50+h100*v100
                control.append({'Agente':a,'Dias_Disponibles':avail[a],'AERO_01_07':counts[a]['AERO_01_07'],'AERO_07_13':counts[a]['AERO_07_13'],'AERO_13_19':counts[a]['AERO_13_19'],'AERO_19_01':counts[a]['AERO_19_01'],'SEC_50':counts[a]['SEC_50'],'SEC_100':counts[a]['SEC_100'],'TOTAL_TURNOS':counts[a]['TOTAL_TURNOS'],'CANT_HORAS_50':h50,'CANT_HORAS_100':h100,'CANT_HORAS_TOTALES':h50+h100,'VALOR_TOTAL_50':h50*v50,'VALOR_TOTAL_100':h100*v100,'VALOR_TOTAL_AGENTE':total,'Sueldo_Proporcional_Por_Dia':round(total/avail[a],2)})
            lrows=[]
            for a in AGENTES:
                l=set(lic.get(a,[]));n=set(no.get(a,[]));ab=set(absences.get(a,[]));ext=set()
                for d in l:
                    nx=d+1
                    if nx<=days and special(nx) and nx not in l:ext.add(nx)
                lrows.append({'Agente':a,'Días Licencia':len(l),'Fechas Licencia':self.ranges(l),'Días Extensión (Finde/Feriado)':len(ext),'Fechas Extensión Finde/Feriado':self.ranges(ext),'Días No Trabajables':len(n),'Fechas Días No Trabajables':self.ranges(n),'AUSENCIAS EXTRAORDINARIAS':len(ab),'Fechas Ausencias':self.ranges(ab),'Total Días No Disponibles':len(l|ext|n|ab)})
            self.result={'Y':Y,'M':M,'days':days,'hol':hol,'v50':v50,'v100':v100,'cron':cron,'df':df,'control':pd.DataFrame(control),'lic':pd.DataFrame(lrows),'abs':absences,'replacements':replacements}
            self.status.set(f'{MESES[M]} {Y}: generado. Ausencias procesadas: {sum(map(len,absences.values()))}.')
            messagebox.showinfo('GIRO CUSTOMS','Cronograma generado y ausencias reemplazadas automáticamente respetando licencias, no trabajables, descanso, turnos y equidad.')
        except Exception as e:messagebox.showerror('No se pudo generar',str(e))
    def preview(self):
        if not self.result:self.generate()
        if not self.result:return
        w=tk.Toplevel(self.root);w.title('Vista previa');w.geometry('1100x650');nb=ttk.Notebook(w);nb.pack(fill='both',expand=True)
        for title,df in [('Cronograma',self.result['df']),('Control de Equidad',self.result['control']),('Licencias y Ausencias',self.result['lic'])]:
            f=ttk.Frame(nb);nb.add(f,text=title);tree=ttk.Treeview(f,columns=list(df.columns),show='headings');tree.pack(fill='both',expand=True)
            for c in df.columns:tree.heading(c,text=c);tree.column(c,width=max(100,min(220,len(c)*9+25)))
            for r in df.itertuples(False):tree.insert('', 'end',values=list(r))
    def export(self):
        if not self.result:self.generate()
        if not self.result:return
        r=self.result;p=filedialog.asksaveasfilename(defaultextension='.xlsx',initialfile=f'Giro_Customs_{r["M"]}_{r["Y"]}.xlsx',filetypes=[('Excel','*.xlsx')])
        if not p:return
        wb=openpyxl.Workbook();ws=wb.active;ws.title='Cronograma Día a Día';self.sheet_df(ws,r['df'],f'CRONOGRAMA - {MESES[r["M"]]} {r["Y"]}')
        mat=wb.create_sheet('Cronograma Consolidado');mat.append([]);mat.append([]);mat.append(['','DIA - HORARIO A CUBRIR','','SECCIÓN AEROPUERTO','','','','ZONA SECUNDARIA','','','','','']);mat.append(['','','','01 A 07','07 A 13','13 A 19','19 A 01','ZONA 1','ZONA 2','ZONA 3','ZONA 4','ZONA 5','ZONA 6'])
        mp=[(4,'AERO_01_07'),(5,'AERO_07_13'),(6,'AERO_13_19'),(7,'AERO_19_01'),(8,'ZONA_1'),(9,'ZONA_2'),(10,'ZONA_3'),(11,'ZONA_4'),(12,'ZONA_5'),(13,'ZONA_6')]
        for d in range(1,r['days']+1):
            dt=date(r['Y'],r['M'],d);row=d+4;mat.cell(row,2,dt.strftime('%A').upper());mat.cell(row,3,d)
            for c,k in mp:mat.cell(row,c,r['cron'][d].get(k,''))
        self.sheet_df(wb.create_sheet('Control de Equidad'),r['control'],'CONTROL DE EQUIDAD');self.sheet_df(wb.create_sheet('Consolidado de Licencias'),r['lic'],'LICENCIAS, NO TRABAJABLES Y AUSENCIAS')
        wb.save(p);self.status.set('Excel guardado: '+p);messagebox.showinfo('Excel','Archivo guardado correctamente.')
    @staticmethod
    def sheet_df(ws,df,title):
        ws['A1']=title;ws['A1'].font=Font(size=14,bold=True,color='1F4E79')
        for c,x in enumerate(df.columns,1):cell=ws.cell(3,c,x);cell.fill=PatternFill('solid',fgColor='1F4E79');cell.font=Font(color='FFFFFF',bold=True);cell.alignment=Alignment(horizontal='center')
        for r,row in enumerate(df.itertuples(False),4):
            for c,v in enumerate(row,1):cell=ws.cell(r,c,v);cell.border=Border(*(Side(style='thin',color='BDC3C7') for _ in range(4)));cell.alignment=Alignment(horizontal='center')
        for col in ws.columns:
            letter=get_column_letter(col[0].column);ws.column_dimensions[letter].width=max(12,min(42,max((len(str(x.value or '')) for x in col),default=10)+4))

if __name__=='__main__':
    root=tk.Tk();GiroCustoms(root);root.mainloop()
